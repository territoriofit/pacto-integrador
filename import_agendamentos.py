# -*- coding: utf-8 -*-
"""
Import da planilha AGENDAMENTO.xlsx (aulas experimentais) -> tabela agendamentos.

Uso:
    python import_agendamentos.py [caminho.xlsx] [--force]

Cada aba (MAIO/JUNHO/JULHO/...) vira um mes_referencia. Idempotente: se o mes
ja tem linhas origem='planilha', pula (use --force para apagar e reimportar).
"""
import re
import sys
import unicodedata
from datetime import date, datetime, time

import openpyxl

from agente_integrador_pacto import CRMClient

XLSX_PADRAO = r"C:\Users\Acer\OneDrive\Desktop\LISTAS CRM\AGENDAMENTO.xlsx"

MES_POR_ABA = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "MARÇO": 3, "ABRIL": 4,
    "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9,
    "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

CONSULTOR_CANONICO = {
    "rai": "Raiane", "raiane": "Raiane",
    "nathy": "Nathalia", "nahy": "Nathalia", "nathalia": "Nathalia",
    "kelytta": "Kellyta", "kellyta": "Kellyta", "ly": "Kellyta",
}

# nomes canonicos das modalidades (mesma lista do frontend: MODALIDADES)
AULA_CANONICA = {
    "musc": "Musculação", "exp musc": "Musculação", "musculacao": "Musculação",
    "visita": "Visita", "viista": "Visita", "visitar": "Visita",
    "fechar": "Fechar", "visita/fechar": "Visita/Fechar",
    "bike": "Bike", "pilates": "Pilates", "funcional": "Funcional",
    "fit dance": "FitDance", "fitdance": "FitDance", "ritmos": "Ritmos",
    "jump": "Jump", "jump/ bike": "Jump", "jump/bike": "Jump",
    "funcional /musc": "Funcional", "funcional/musc": "Funcional",
}


def _norm(s) -> str:
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def parse_consultor(v) -> str | None:
    n = _norm(v)
    if not n:
        return None
    # 'nathy /ly', 'nathy,' -> primeiro nome reconhecivel
    for token in re.split(r"[\s,/]+", n):
        if token in CONSULTOR_CANONICO:
            return CONSULTOR_CANONICO[token]
    return str(v).strip().title()


def parse_fonte(v) -> str | None:
    n = _norm(v)
    if not n:
        return None
    if n.startswith(("lead", "leed", "lrad")):
        return "Leads"
    if "indica" in n:
        return "Indicação"
    if "liga" in n:
        return "Ligação"
    if "balcao" in n:
        return "Balcão"
    if "instagram" in n:
        return "Instagram"
    if n.isdigit():  # matricula de quem indicou
        return "Indicação"
    return str(v).strip().capitalize()


def parse_aula(v) -> str | None:
    n = _norm(v)
    if not n:
        return None
    return AULA_CANONICA.get(n, str(v).strip().capitalize())


def parse_bool(v):
    n = _norm(v)
    if n in ("sim", "s"):
        return True
    if n in ("nao", "n"):
        return False
    return None


def parse_data(v, ano: int, mes: int):
    """datetime da celula ou texto tipo '04/05/' / '01/06.' -> date ISO."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{1,2})", str(v or ""))
    if m:
        d, mm = int(m.group(1)), int(m.group(2))
        try:
            return date(ano, mm, d).isoformat()
        except ValueError:
            return None
    return None


def parse_horario(v):
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    s = str(v or "").strip()
    return s or None


def parse_telefone(v):
    dig = re.sub(r"\D", "", str(v or ""))
    return dig or None


def importar(xlsx: str, force: bool = False):
    crm = CRMClient()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ano = date.today().year
    total = 0

    for ws in wb.worksheets:
        mes_num = MES_POR_ABA.get(_norm(ws.title).upper())
        if not mes_num:
            print(f"[{ws.title}] aba ignorada (nao e mes)")
            continue
        mes_ref = f"{ano}-{mes_num:02d}"

        existentes = crm.sb.table("agendamentos").select("id", count="exact") \
            .eq("mes_referencia", mes_ref).eq("origem", "planilha").limit(1).execute()
        if (existentes.count or 0) > 0:
            if not force:
                print(f"[{ws.title}] {mes_ref} ja importado ({existentes.count} linhas) — pulando (use --force)")
                continue
            crm.sb.table("agendamentos").delete() \
                .eq("mes_referencia", mes_ref).eq("origem", "planilha").execute()
            print(f"[{ws.title}] {mes_ref} reimportando (--force)")

        # header: linha cujo col D == 'nome'
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            if _norm(row[3]) == "nome":
                header_row = i
                break
        if not header_row:
            print(f"[{ws.title}] header nao encontrado — pulando")
            continue

        linhas = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            nome = str(row[3] or "").strip()
            if not nome:
                continue
            horario = parse_horario(row[8])
            confirmacao = parse_bool(row[9])
            # linhas onde a vendedora pulou a coluna horario e o 'sim' caiu nela
            if confirmacao is None and _norm(horario) in ("sim", "nao", "não"):
                confirmacao = parse_bool(horario)
                horario = None
            linhas.append({
                "nome":             nome,
                "telefone":         parse_telefone(row[5]),
                "fonte":            parse_fonte(row[4]),
                "aula":             parse_aula(row[6]),
                "consultor":        parse_consultor(row[7]),
                "mes_referencia":   mes_ref,
                "semana":           row[0] if isinstance(row[0], int) else None,
                "data_contato":     parse_data(row[1], ano, mes_num),
                "data_agendamento": parse_data(row[2], ano, mes_num),
                "horario":          horario,
                "confirmacao":      confirmacao,
                "veio":             parse_bool(row[10]),
                "fechou":           parse_bool(row[11]),
                "observacao":       (str(row[12]).strip() if row[12] else None),
                "origem":           "planilha",
                "tenant_id":        crm.tenant_id,
            })

        # vincula lead pelo final do telefone (8 digitos); ambiguo = sem vinculo
        for ln in linhas:
            tel = ln["telefone"]
            if not tel or len(tel) < 8:
                continue
            r = crm.sb.table("leads").select("id").like("phone", f"%{tel[-8:]}").limit(2).execute()
            if r.data and len(r.data) == 1:
                ln["lead_id"] = r.data[0]["id"]

        for i in range(0, len(linhas), 100):
            crm.sb.table("agendamentos").insert(linhas[i:i + 100]).execute()
        com_lead = sum(1 for x in linhas if x.get("lead_id"))
        print(f"[{ws.title}] {mes_ref}: {len(linhas)} agendamentos importados ({com_lead} vinculados a lead)")
        total += len(linhas)

    print(f"Total: {total} linhas")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--force"]
    importar(args[0] if args else XLSX_PADRAO, force="--force" in sys.argv)
