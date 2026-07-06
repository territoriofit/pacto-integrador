# -*- coding: utf-8 -*-
"""
Import da planilha leads.xlsx (acompanhamento de leads) -> tabela leads_acompanhamento.

Uso:
    python import_leads_acompanhamento.py [caminho.xlsx] [--force]

As 10 colunas "Tentativa N" viram um unico ultimo_contato (data mais recente).
mes_referencia = mes predominante da coluna Data (planilha e mensal).
Idempotente por mes (--force apaga e reimporta as linhas origem='planilha').
"""
import re
import sys
import unicodedata
from collections import Counter
from datetime import date, datetime

import openpyxl

from agente_integrador_pacto import CRMClient

XLSX_PADRAO = r"C:\Users\Acer\OneDrive\Desktop\LISTAS CRM\leads.xlsx"

MESES_PT = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
            "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}

CONSULTORA_CANONICA = {
    "rai": "Raiane", "raiane": "Raiane",
    "nathy": "Nathalia", "nathalia": "Nathalia", "nahy": "Nathalia",
    "kelytta": "Kellyta", "kellyta": "Kellyta", "kelly": "Kellyta", "ly": "Kellyta",
}


def _norm(s) -> str:
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def parse_bool(v):
    n = _norm(v)
    if n.startswith("s"):
        return True
    if n.startswith("n"):
        return False
    return None


def parse_tipo(v):
    n = _norm(v)
    if n.startswith("patr"):
        return "patrocinado"
    if n.startswith("espo") or n.startswith("esponat"):
        return "espontaneo"
    return None


def parse_telefone(v):
    if isinstance(v, (int, float)):
        v = f"{int(v)}"
    dig = re.sub(r"\D", "", str(v or ""))
    return dig or None


def parse_data(v, ano: int):
    """datetime da celula ou texto 'DD/mmm' -> date."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r"^\s*(\d{1,2})\s*/\s*([a-z]{3})", _norm(v))
    if m:
        mes = MESES_PT.get(m.group(2))
        if mes:
            try:
                return date(ano, mes, int(m.group(1)))
            except ValueError:
                return None
    return None


def importar(xlsx: str, force: bool = False):
    crm = CRMClient()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.worksheets[0]
    ano = date.today().year

    # header: linha com 'Nome' na coluna D
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if _norm(row[3]) == "nome":
            header_row = i
            break
    if not header_row:
        print("header nao encontrado")
        return

    brutas = [row for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                                          values_only=True)
              if row[3] or row[4]]  # nome OU numero (planilha tem linhas sem nome)

    # mes predominante da coluna Data define o mes_referencia da planilha
    meses = Counter(d.month for d in (parse_data(r[1], ano) for r in brutas) if d)
    mes_num = meses.most_common(1)[0][0]
    mes_ref = f"{ano}-{mes_num:02d}"

    existentes = crm.sb.table("leads_acompanhamento").select("id", count="exact") \
        .eq("mes_referencia", mes_ref).eq("origem", "planilha").limit(1).execute()
    if (existentes.count or 0) > 0:
        if not force:
            print(f"{mes_ref} ja importado ({existentes.count}+ linhas) — use --force")
            return
        crm.sb.table("leads_acompanhamento").delete() \
            .eq("mes_referencia", mes_ref).eq("origem", "planilha").execute()
        print(f"{mes_ref} reimportando (--force)")

    linhas = []
    for row in brutas:
        telefone = parse_telefone(row[4])
        # ultimo contato = data mais recente entre Tentativa 1-10 e ULTIMO CONTATO
        datas = [parse_data(row[c], ano) for c in range(12, 23) if len(row) > c]
        datas = [d for d in datas if d]
        linhas.append({
            "nome":           (str(row[3]).strip() if row[3] else None),
            "telefone":       telefone,
            "consultora":     CONSULTORA_CANONICA.get(_norm(row[2]),
                                                      (str(row[2]).strip().title() if row[2] else None)),
            "mes_referencia": mes_ref,
            "semana":         row[0] if isinstance(row[0], int) else None,
            "data":           (parse_data(row[1], ano) or date(ano, mes_num, 1)).isoformat(),
            "fonte":          (str(row[5]).strip()[:300] if row[5] else None),
            "tipo":           parse_tipo(row[6]),
            "agendou":        parse_bool(row[7]),
            "veio":           parse_bool(row[8]),
            "fechou":         parse_bool(row[9]),
            "objecao":        (str(row[10]).strip() if row[10] else None),
            "ultima_msg":     (str(row[11]).strip() if row[11] else None),
            "ultimo_contato": max(datas).isoformat() if datas else None,
            "origem":         "planilha",
            "tenant_id":      crm.tenant_id,
        })

    # vincula lead pelo final do telefone (8 digitos); ambiguo = sem vinculo
    for ln in linhas:
        tel = ln["telefone"]
        if not tel or len(tel) < 8:
            continue
        r = crm.sb.table("leads").select("id").like("phone", f"%{tel[-8:]}").limit(2).execute()
        if r.data and len(r.data) == 1:
            ln["lead_id"] = r.data[0]["id"]

    for i in range(0, len(linhas), 100):
        crm.sb.table("leads_acompanhamento").insert(linhas[i:i + 100]).execute()
    com_lead = sum(1 for x in linhas if x.get("lead_id"))
    print(f"{mes_ref}: {len(linhas)} leads importados ({com_lead} vinculados a lead)")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--force"]
    importar(args[0] if args else XLSX_PADRAO, force="--force" in sys.argv)
