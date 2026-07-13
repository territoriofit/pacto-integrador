"""
Agente Integrador Pacto -- Territorio Fit
Modulos: Alunos, Frequencia (ZW), Contratos, Notificacoes
"""

import os
import json
import time
import uuid
import logging
import requests
from datetime import datetime, timedelta, date, timezone as dt_timezone
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env.integracao")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

EXPORT_DIR = Path(r"G:\Meu Drive\André Obsidian\André Obsidian\09 - Dados Pacto")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

PACTO_BASE_URL = os.getenv("PACTO_BASE_URL", "https://zw819.pactosolucoes.com.br/TreinoWeb/prest")
PACTO_TOKEN    = os.getenv("PACTO_TOKEN")
PACTO_CHAVE    = os.getenv("PACTO_CHAVE")
PACTO_EMP_ID   = os.getenv("PACTO_EMPRESA_ID")
PACTO_CTX      = os.getenv("PACTO_CTX")
PACTO_USER     = os.getenv("PACTO_USER")
PACTO_PWD      = os.getenv("PACTO_PWD")


def _ts_to_date(ts) -> datetime | None:
    """Converte timestamp Unix em ms para datetime. Retorna None se invalido."""
    if ts is None:
        return None
    try:
        return datetime.fromtimestamp(int(ts) / 1000)
    except Exception:
        return None


class PactoClient:
    """Cliente HTTP para a API TreinoWeb da Pacto Solucoes -- Territorio Fit."""

    def __init__(self):
        self.base        = PACTO_BASE_URL
        self.ctx         = PACTO_CTX
        self.jwt_token   = None
        self._authed     = False
        self.session     = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {PACTO_TOKEN}",
            "chave":         PACTO_CHAVE,
            "empresaId":     PACTO_EMP_ID,
            "Content-Type":  "application/json",
        })
        # autenticacao automatica na inicializacao
        if not self.login_jwt():
            log.warning("Nao foi possivel autenticar automaticamente -- usando token estatico")

    # -- Requisicao com retry + re-auth automatico ----------------------------

    def _req(self, method: str, path: str, _retry_auth: bool = True, **kwargs):
        url = f"{self.base}{path}"
        for tentativa in range(3):
            try:
                r = self.session.request(method, url, timeout=30, **kwargs)
                if r.status_code in (200, 201):
                    try:
                        return r.json()
                    except Exception:
                        return {"raw": r.text}
                elif r.status_code == 401:
                    if _retry_auth:
                        log.info("Token expirado -- renovando autenticacao...")
                        if self.login_jwt():
                            return self._req(method, path, _retry_auth=False, **kwargs)
                    raise Exception("Token invalido ou expirado apos renovacao")
                elif r.status_code == 403:
                    raise Exception(f"Sem permissao: {path}")
                elif r.status_code >= 500:
                    log.warning(f"Pacto 500 em {path} -- tentativa {tentativa+1}/3")
                    time.sleep(5)
                else:
                    log.error(f"Pacto {r.status_code} em {path}: {r.text[:200]}")
                    return {"erro": r.status_code, "mensagem": r.text[:200]}
            except requests.exceptions.ConnectionError:
                log.warning(f"Falha de conexao -- tentativa {tentativa+1}/3")
                time.sleep(3)
        raise Exception(f"Maximo de tentativas atingido: {path}")

    # -- Autenticacao JWT ------------------------------------------------------

    def login_jwt(self) -> bool:
        """
        Obtem token JWT e reconfigura session com empresaId=1.
        O empresaId '1' e o indice da Territorio Fit no payload do JWT.
        """
        try:
            r = requests.post(
                f"{self.base}/login",
                json={"username": PACTO_USER, "senha": PACTO_PWD, "chave": PACTO_CHAVE},
                headers={"Content-Type": "application/json"},
                timeout=15,
            )
            if r.status_code == 200:
                data = r.json()
                self.jwt_token = (
                    data.get("token") or data.get("access_token")
                    or data.get("jwt") or data.get("content")
                )
                if self.jwt_token:
                    self.session.headers.update({
                        "Authorization": f"Bearer {self.jwt_token}",
                        "empresaId": "1",
                    })
                    log.info("Login JWT realizado com sucesso")
                    return True
            log.error(f"Falha no login JWT: {r.status_code} -- {r.text[:200]}")
        except Exception as e:
            log.error(f"Erro no login JWT: {e}")
        return False

    # -- Descoberta de endpoints -----------------------------------------------

    def listar_empresas(self) -> dict:
        return self._req("GET", "/EndpointControl/empresas")

    def listar_endpoints(self) -> dict:
        return self._req("GET", "/EndpointControl/show")

    def views_bi(self) -> dict:
        return self._req("GET", "/psec/grafico-bi/views/all")

    # -- ALUNOS ----------------------------------------------------------------

    def todos_alunos(self, situacao: str = None) -> list:
        """
        Retorna alunos via /psec/alunos (requer JWT + empresaId=1).
        situacao: 'ATIVO', 'INATIVO', 'VISITANTE' ou None para todos.

        Campos principais:
          id, matriculaZW, nome, situacaoAluno, dataNascimento, sexo,
          emails, fones, planoZW, contratoZW {vencimento (timestamp ms), tipo},
          codigoCliente, codigoPessoa, programas, parq_status.
        """
        todos       = []
        page        = 0
        size        = 500
        total_pages = None
        while True:
            r = self._req("GET", "/psec/alunos", params={"size": size, "page": page})
            content = r.get("content", []) if isinstance(r, dict) else (r or [])
            if not content:
                break
            if total_pages is None:
                total_pages = r.get("totalPages", 1)
            if situacao:
                content = [a for a in content if a.get("situacaoAluno") == situacao]
            todos.extend(content)
            page += 1
            if page >= total_pages:
                break
        return todos

    def alunos_ativos(self) -> list:
        resultado = self.todos_alunos(situacao="ATIVO")
        log.info(f"Alunos ativos: {len(resultado)}")
        return resultado

    def alunos_inativos(self) -> list:
        resultado = self.todos_alunos(situacao="INATIVO")
        log.info(f"Alunos inativos (ex-alunos): {len(resultado)}")
        return resultado

    def checkins_recentes(self) -> list:
        """Ultimos check-ins agregados da academia (todos os alunos)."""
        r = self._req("GET", "/psec/alunos/ultimos-checkins-agregados")
        return r.get("content", []) if isinstance(r, dict) else (r or [])

    def lista_acessos_recentes(self, tipo: int = 1, limite: int = 50) -> list:
        """
        Lista rapida dos ultimos acessos fisicos registrados na academia.
        tipo=1 lista todos; tipo=2 pode filtrar por tipo especifico (enum da API).
        Retorna campos: codigoCliente, matricula, hora, nome, plano, situacao, tipoCheckin.
        """
        r = self._req("GET", "/psec/alunos/lista-rapida-acessos",
                      params={"tipo": tipo, "limite": limite})
        content = r.get("content", {}) if isinstance(r, dict) else {}
        return content.get("lista", []) if isinstance(content, dict) else []

    def historico_presenca(self, matricula: int) -> dict:
        """
        Resumo de presenca do aluno em aulas/turmas agendadas:
          totalAulasRealizadas, aulasMesAtual, semanasConsecutivas.
        Nota: conta apenas aulas agendadas pelo sistema -- nao conta treino livre na catraca.
        Requer param empresa=1.
        """
        r = self._req("GET", f"/cliente/{self.ctx}/historico-presenca",
                      params={"matricula": matricula, "empresa": 1, "atualizaCache": True})
        return r.get("content", {}) if isinstance(r, dict) else {}

    def distribuicao_acesso_semanal(self, cod_aluno: int,
                                    data_ini: date = None, data_fim: date = None) -> dict:
        """
        Distribuicao de acessos fisicos do aluno por dia da semana no periodo.
        Usa dados de catraca/check-in fisico (nao TreinoWeb app).
        Retorna dict com chaves SEG, TER, QUA, QUI, SEX, SAB, DOM.
        data_ini/data_fim default: ultimos 6 meses.
        """
        if data_fim is None:
            data_fim = date.today()
        if data_ini is None:
            data_ini = date(data_fim.year - (1 if data_fim.month <= 6 else 0),
                            (data_fim.month - 6) % 12 or 12, 1)
        ts_ini = int(datetime(data_ini.year, data_ini.month, data_ini.day).timestamp() * 1000)
        ts_fim = int(datetime(data_fim.year, data_fim.month, data_fim.day).timestamp() * 1000)
        r = self._req("GET",
                      f"/cliente/{self.ctx}/app/consultarQuantidadeAcessosClientesAgrupadosDia",
                      params={"codigoAluno": cod_aluno, "username": "",
                              "dataInicial": ts_ini, "dataFinal": ts_fim})
        if isinstance(r, dict) and "sucesso" in r:
            return r["sucesso"]
        return {}

    def linha_tempo_aluno(self, matricula: int) -> list:
        """
        Historico de eventos do aluno no sistema ZW:
        ACABOU_TREINO, MONTOU_TREINO, REVISOU_TREINO, checkins, etc.
        O campo 'data' e timestamp Unix em ms.
        """
        r = self._req("GET", f"/psec/alunos/linha-tempo/{matricula}")
        return r.get("content", []) if isinstance(r, dict) else (r or [])

    def ultimo_treino(self, matricula: int) -> datetime | None:
        """Retorna a data do ultimo treino concluido (ACABOU_TREINO) do aluno."""
        eventos = self.linha_tempo_aluno(matricula)
        for ev in eventos:
            if ev.get("evento") in ("ACABOU_TREINO", "CHECKIN", "ACESSOU"):
                return _ts_to_date(ev.get("data"))
        return None

    def ultimo_acesso_catraca(self, matricula: int) -> datetime | None:
        """
        Data do ultimo acesso fisico (catraca) do aluno, via linha-tempo.
        O acesso na catraca gera evento NOTIFICACAO cuja descricao varia:
        'Notificacao: Chegou' ou 'Notificacao: Aluno chegou e possui Indice
        de Risco' (grupo de risco) — por isso o match e case-insensitive em
        'chegou'. Cobre tambem check-ins Gympass/TotalPass.
        """
        eventos = self.linha_tempo_aluno(matricula)
        for ev in eventos:
            desc = (ev.get("descricao") or "").lower()
            if "chegou" in desc or ev.get("evento") in ("CHECKIN", "ACESSOU"):
                return _ts_to_date(ev.get("data"))
        return None

    def alunos_faltosos(self, dias: int = 7) -> list:
        """
        Retorna alunos ATIVOS sem treino registrado nos ultimos `dias` dias.
        Usa linha-tempo por aluno (evento ACABOU_TREINO) -- sistema ZW musculacao.
        """
        ativos   = self.alunos_ativos()
        corte    = datetime.now() - timedelta(days=dias)
        faltosos = []

        log.info(f"Verificando {len(ativos)} alunos ativos (ultimos {dias} dias)...")

        for aluno in ativos:
            matricula = aluno.get("matriculaZW") or aluno.get("id")
            if not matricula:
                continue
            try:
                ultimo = self.ultimo_treino(matricula)
                if ultimo is None or ultimo < corte:
                    dias_aus = (datetime.now() - ultimo).days if ultimo else 999
                    faltosos.append({
                        **aluno,
                        "ultimo_treino": ultimo.strftime("%d/%m/%Y") if ultimo else None,
                        "dias_ausente":  dias_aus,
                    })
            except Exception as e:
                log.debug(f"Erro linha-tempo aluno {matricula}: {e}")

        faltosos.sort(key=lambda x: x.get("dias_ausente") or 999, reverse=True)
        log.info(f"Alunos faltosos (>{dias} dias): {len(faltosos)}")
        return faltosos

    # -- CONTRATOS / FINANCEIRO ------------------------------------------------

    def inadimplentes(self) -> list:
        """
        Retorna alunos ATIVOS com contrato vencido.
        O campo contratoZW.vencimento e timestamp Unix em ms.
        """
        ativos = self.alunos_ativos()
        hoje   = datetime.now()
        inadimpl = []
        for aluno in ativos:
            ts = (aluno.get("contratoZW") or {}).get("vencimento")
            venc = _ts_to_date(ts)
            if venc and venc < hoje:
                inadimpl.append({
                    **aluno,
                    "vencimento_fmt": venc.strftime("%d/%m/%Y"),
                    "dias_atraso":    (hoje - venc).days,
                })
        inadimpl.sort(key=lambda x: x.get("dias_atraso", 0), reverse=True)
        log.info(f"Inadimplentes (contrato vencido): {len(inadimpl)}")
        return inadimpl

    def cancelamentos(self, dias: int = 30) -> list:
        """
        Retorna alunos INATIVO como representacao de cancelamentos.
        Filtra pelos que tem contratoZW.vencimento nos ultimos `dias` dias.
        """
        inativos = self.alunos_inativos()
        corte    = datetime.now() - timedelta(days=dias)
        recentes = []
        for aluno in inativos:
            ts   = (aluno.get("contratoZW") or {}).get("vencimento")
            venc = _ts_to_date(ts)
            if venc and venc >= corte:
                recentes.append({**aluno, "data_cancelamento": venc.strftime("%d/%m/%Y")})
        log.info(f"Cancelamentos (ultimos {dias} dias): {len(recentes)}")
        return recentes

    def renovacoes(self, dias: int = 30) -> list:
        """
        Retorna alunos ATIVOS com contratoZW.tipo == RENOVACAO cujo vencimento
        e nos proximos `dias` dias (contratos recem-renovados com vencimento futuro).
        """
        ativos  = self.alunos_ativos()
        corte   = datetime.now() - timedelta(days=dias)
        hoje    = datetime.now()
        renovados = []
        for aluno in ativos:
            contrato = aluno.get("contratoZW") or {}
            if contrato.get("tipo") in ("RENOVACAO", "REMATRICULA"):
                ts   = contrato.get("vencimento")
                venc = _ts_to_date(ts)
                if venc and venc > hoje:
                    renovados.append({**aluno, "vencimento_fmt": venc.strftime("%d/%m/%Y")})
        log.info(f"Renovacoes ativas: {len(renovados)}")
        return renovados

    def matriculas_novas(self, dias: int = 30) -> list:
        """
        Retorna alunos ATIVOS com contratoZW.tipo == MATRICULA com vencimento
        nos proximos `dias` * 12 dias (matriculas recentes ainda vigentes).
        """
        ativos = self.alunos_ativos()
        corte  = datetime.now() - timedelta(days=dias)
        hoje   = datetime.now()
        novas  = []
        for aluno in ativos:
            contrato = aluno.get("contratoZW") or {}
            if contrato.get("tipo") == "MATRICULA":
                ts   = contrato.get("vencimento")
                venc = _ts_to_date(ts)
                if venc and venc > hoje:
                    novas.append({**aluno, "vencimento_fmt": venc.strftime("%d/%m/%Y")})
        log.info(f"Matriculas ativas (tipo MATRICULA): {len(novas)}")
        return novas

    def contratos_a_vencer(self, dias: int = 30) -> list:
        """Alunos ATIVOS com contrato vencendo nos proximos `dias` dias."""
        ativos   = self.alunos_ativos()
        hoje     = datetime.now()
        limite   = hoje + timedelta(days=dias)
        a_vencer = []
        for aluno in ativos:
            ts   = (aluno.get("contratoZW") or {}).get("vencimento")
            venc = _ts_to_date(ts)
            if venc and hoje <= venc <= limite:
                a_vencer.append({
                    **aluno,
                    "vencimento_fmt": venc.strftime("%d/%m/%Y"),
                    "dias_restantes": (venc - hoje).days,
                })
        a_vencer.sort(key=lambda x: x.get("dias_restantes", 0))
        log.info(f"Contratos a vencer em {dias} dias: {len(a_vencer)}")
        return a_vencer

    # -- AGENDAMENTOS ----------------------------------------------------------

    def historico_agendamentos(self, codigo_cliente: int, max_results: int = 10, index: int = 0) -> dict:
        return self._req("POST", f"/agendamento/{self.ctx}/historico",
                         params={"codigoCliente": codigo_cliente,
                                 "maxResults": max_results, "index": index})

    def confirmar_agendamento(self, username: str, id_agendamento: str, matricula: int) -> dict:
        return self._req("POST", f"/agendamento/{self.ctx}/confirma",
                         params={"username": username, "idAgendamento": id_agendamento,
                                 "matricula": matricula})

    def cancelar_agendamento(self, username: str, id_agendamento: str, matricula: int) -> dict:
        return self._req("POST", f"/agendamento/{self.ctx}/cancelar",
                         params={"username": username, "idAgendamento": id_agendamento,
                                 "matricula": matricula})

    # -- NOTIFICACOES ----------------------------------------------------------

    def push_lembrar_compromisso(self, data_atual: str) -> dict:
        return self._req("POST", f"/EndpointControl/{self.ctx}/pushLembrarCompromisso",
                         params={"dataAtual": data_atual})

    def push_proximos_agendados(self, data_atual: str, tipo_lembrete: str = "QUINZE_MINUTOS") -> dict:
        return self._req("POST", f"/EndpointControl/{self.ctx}/pushProximosAgendados",
                         params={"dataAtual": data_atual, "tipoLembrete": tipo_lembrete})

    # -- WHATSAPP (stub -- conectar ao novo CRM) --------------------------------

    def notificar_whatsapp(self, telefone: str, mensagem: str, crm_client=None) -> bool:
        """
        Envia mensagem WhatsApp via CRM.
        Passar `crm_client` com metodo .enviar_whatsapp(tel, msg) quando CRM estiver definido.
        """
        if crm_client is None:
            log.info(f"[WhatsApp STUB] Para: {telefone} | Msg: {mensagem[:60]}...")
            return True
        return crm_client.enviar_whatsapp(telefone, mensagem)

    # -- USUÁRIO / ADM ---------------------------------------------------------

    def obter_usuario(self) -> dict:
        """
        Dados do usuário autenticado: nome, perfil, recursos e funcionalidades.
        Usa /psec/validateToken (TreinoWeb) que retorna JSON em claro.
        O /adm/obter-usuario (API Gateway) retorna AES-encriptado -- chave desconhecida.
        """
        r = self._req("GET", "/psec/validateToken")
        return r.get("content", r) if isinstance(r, dict) else r

    # -- DASHBOARD CONSOLIDADO -------------------------------------------------

    def dashboard(self) -> dict:
        """Gera snapshot consolidado da operacao."""
        log.info("Gerando dashboard...")
        agora = datetime.now()

        ativos      = self.alunos_ativos()
        inadimpl    = self.inadimplentes()
        a_vencer    = self.contratos_a_vencer(dias=30)
        cancelados  = self.cancelamentos(dias=30)
        faltosos_7  = self.alunos_faltosos(dias=7)
        faltosos_14 = self.alunos_faltosos(dias=14)

        return {
            "gerado_em": agora.strftime("%d/%m/%Y %H:%M"),
            "alunos": {
                "ativos":           len(ativos),
                "faltosos_7_dias":  len(faltosos_7),
                "faltosos_14_dias": len(faltosos_14),
            },
            "contratos": {
                "inadimplentes":      len(inadimpl),
                "a_vencer_30_dias":   len(a_vencer),
                "cancelamentos_30d":  len(cancelados),
            },
            "detalhes": {
                "faltosos_7_dias": faltosos_7[:20],
                "inadimplentes":   inadimpl[:20],
                "a_vencer":        a_vencer[:20],
                "cancelamentos":   cancelados[:10],
            },
        }

    # -- EXPORTACAO ------------------------------------------------------------

    def exportar(self, nome: str, dados) -> Path:
        arquivo = EXPORT_DIR / f"{nome}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        with open(arquivo, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        log.info(f"Exportado: {arquivo}")
        return arquivo

    def exportar_dashboard_md(self, painel: dict) -> Path:
        a = painel["alunos"]
        c = painel["contratos"]

        linhas_faltosos = ""
        for al in painel["detalhes"]["faltosos_7_dias"]:
            fones = al.get("fones") or []
            tel   = fones[0].get("numero") if fones else "—"
            linhas_faltosos += (
                f"- **{al.get('nome', '—')}** | Tel: {tel} "
                f"| Ausente: {al.get('dias_ausente', '—')} dias "
                f"| Ultimo treino: {al.get('ultimo_treino') or 'nenhum'}\n"
            )

        linhas_inadimpl = ""
        for al in painel["detalhes"]["inadimplentes"]:
            linhas_inadimpl += (
                f"- **{al.get('nome', '—')}** "
                f"| Venceu: {al.get('vencimento_fmt', '—')} "
                f"| Atraso: {al.get('dias_atraso', '—')} dias\n"
            )

        md = f"""# Dashboard Operacional — {painel['gerado_em']}

## Alunos
| Metrica | Valor |
|---|---|
| Ativos | {a['ativos']} |
| Faltosos +7 dias | {a['faltosos_7_dias']} |
| Faltosos +14 dias | {a['faltosos_14_dias']} |

## Contratos
| Metrica | Valor |
|---|---|
| Inadimplentes | {c['inadimplentes']} |
| A vencer em 30 dias | {c['a_vencer_30_dias']} |
| Cancelamentos (30d) | {c['cancelamentos_30d']} |

## Alunos faltosos +7 dias
{linhas_faltosos}
## Inadimplentes
{linhas_inadimpl}
"""
        arquivo = EXPORT_DIR / f"dashboard_{datetime.now().strftime('%Y%m%d')}.md"
        with open(arquivo, "w", encoding="utf-8") as f:
            f.write(md)
        log.info(f"Dashboard MD: {arquivo}")
        return arquivo


# -- MÓDULO ADM / CRM (API Gateway) ------------------------------------------

SUPABASE_URL   = os.getenv("SUPABASE_URL", "https://bmnyhaxvlifmwkcuglfh.supabase.co")
SUPABASE_KEY   = os.getenv("SUPABASE_KEY")
CRM_TENANT_ID  = os.getenv("CRM_TENANT_ID")

PACTO_GW_URL = "https://apigw.pactosolucoes.com.br"


class PactoADMClient:
    """
    Acesso ao módulo ADM/CRM via API Gateway (apigw.pactosolucoes.com.br).
    Autentica com PACTO_TOKEN (hex) -- header Authorization sem prefixo Bearer.
    """

    def __init__(self):
        self.base    = PACTO_GW_URL
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": PACTO_TOKEN,
            "chave":         PACTO_CHAVE,
            "empresaId":     "1",
            "Content-Type":  "application/json",
        })

    def _gw(self, method: str, path: str, timeout: int = 20, **kwargs):
        r = self.session.request(method, f"{self.base}{path}", timeout=timeout, **kwargs)
        if r.status_code in (200, 201):
            try:
                return r.json()
            except Exception:
                return {"raw": r.text}
        log.error(f"ADM GW {r.status_code} em {path}: {r.text[:200]}")
        return {"erro": r.status_code, "mensagem": r.text[:200]}

    def _epoch_ms(self, d: date) -> int:
        from datetime import timezone
        return int(datetime(d.year, d.month, d.day, tzinfo=timezone.utc).timestamp() * 1000)

    def meta_crm(self, data_inicio: date, data_fim: date) -> list:
        """Retorna metas CRM por fase para o período informado."""
        # janelas de 30 dias podem demorar >20s no gateway — timeout maior
        r = self._gw("GET", "/meta-crm", timeout=90, params={
            "dataInicio": self._epoch_ms(data_inicio),
            "dataFim":    self._epoch_ms(data_fim),
        })
        return r.get("content", []) if isinstance(r, dict) else []

    def meta_crm_detalhada(self, codigos: list[int], page: int = 0, size: int = 100) -> dict:
        """Retorna clientes detalhados para os códigos de fechamento de meta."""
        if not codigos:
            return {"content": [], "totalElements": 0}
        params = [("codigosFecharMeta", ",".join(str(c) for c in codigos)),
                  ("page", page), ("size", size)]
        return self._gw("GET", "/meta-crm/detalhada", timeout=90, params=params)

    def visitantes_24h(self, d: date) -> dict:
        """
        Retorna visitantes cadastrados para contato na data informada.
        Combina meta-crm (fase HO) + meta-crm/detalhada para obter nomes e responsáveis.

        Retorna dict com:
          total      -- total_elements da fase HO
          visitantes -- lista com nome, matricula, situacao, dataMeta, consultora
        """
        fases = self.meta_crm(d, d + timedelta(days=1))
        ho_codes = []
        ho_meta = ho_realizado = 0
        for fase in fases:
            for tipo in fase.get("tiposMeta", []):
                if tipo.get("faseEnum") == "HO":
                    ho_codes      = tipo.get("codigosFecharMeta", [])
                    ho_meta       = tipo.get("totalMeta", 0)
                    ho_realizado  = tipo.get("totalMetaRealizada", 0)

        if not ho_codes:
            return {"total": 0, "meta": 0, "realizado": 0, "visitantes": []}

        detalhado = self.meta_crm_detalhada(ho_codes)
        visitantes = [
            {
                "nome":          v.get("nome"),
                "matricula":     v.get("matricula"),
                "codigoCliente": v.get("codigoCliente"),
                "situacao":      v.get("situacao"),
                "dataMeta":      v.get("dataMeta"),
                "consultora":    v.get("nomeColaborador"),
                "telefone":      v.get("telefone"),
                "email":         v.get("email"),
            }
            for v in detalhado.get("content", [])
        ]
        return {
            "total":      detalhado.get("totalElements", len(visitantes)),
            "meta":       ho_meta,
            "realizado":  ho_realizado,
            "visitantes": visitantes,
        }

    def historico_contato(self, page: int = 0, size: int = 100) -> dict:
        """Histórico de contatos CRM (todas as entradas, sem filtro de data)."""
        return self._gw("GET", "/historico-contato", params={"page": page, "size": size})

    def meta_crm_abertura(self) -> dict:
        """
        Verifica se a meta CRM foi aberta hoje e se pode ser aberta.
        Campos: abriuMetaHoje (bool), metaPodeSerAberta (bool), motivo, mensagemBloqueio.
        """
        r = self._gw("GET", "/meta-crm/abertura")
        return r.get("content", r) if isinstance(r, dict) else r

    def fases_crm(self) -> list:
        """
        Fases do funil CRM configuradas.
        Campos por fase: name (enum), descricao, cor, ordem.
        Exemplos: AGENDAMENTO, NEGOCIACAO, MATRICULA, etc.
        """
        r = self._gw("GET", "/v1/configuracao/fases")
        return r.get("result", []) if isinstance(r, dict) else (r or [])

    def fases_crm_ai(self) -> list:
        """Fases do CRM com categorização por IA (NOVOS_LEADS, RETENCAO, etc.)."""
        r = self._gw("GET", "/v1/configuracao/fases-ai")
        return r.get("result", []) if isinstance(r, dict) else (r or [])

    def objecoes(self) -> list:
        """
        Objeções cadastradas para uso no CRM (motivos de perda de oportunidade).
        Campos: codigo, descricao, ativo.
        """
        r = self._gw("GET", "/v1/comum/objecao")
        content = r.get("content", r) if isinstance(r, dict) else r
        return content if isinstance(content, list) else []

    def meta_diaria(self, data_inicio: date, data_fim: date) -> list:
        """
        Meta diária do CRM por colaborador para o período.
        Parâmetros: dataInicio e dataFim no formato ISO 'yyyy-MM-dd'.
        Campos por entrada: colaborador, metaDiaria, realizadoDiario, etc.
        """
        r = self._gw("GET", "/meta-diaria", params={
            "dataInicio": data_inicio.isoformat(),
            "dataFim":    data_fim.isoformat(),
        })
        return r.get("content", []) if isinstance(r, dict) else (r or [])

    # -- ADM (ADMINISTRAÇÃO) --------------------------------------------------

    @staticmethod
    def _decrypt_aes(data: str) -> dict:
        """
        Decripta resposta AES-128 retornada pelos endpoints /adm/*.
        Tenta ECB primeiro, depois CBC com IV zero.
        Chave: PACTO_CHAVE decodificada de hex (16 bytes).
        Requer: pip install pycryptodome
        """
        import base64
        try:
            from Crypto.Cipher import AES
            from Crypto.Util.Padding import unpad
        except ImportError:
            log.warning(
                "pycryptodome nao instalado -- retornando dado encriptado bruto. "
                "Instale com: pip install pycryptodome"
            )
            return {"raw_encrypted": data}

        key = bytes.fromhex(PACTO_CHAVE)
        try:
            raw = base64.b64decode(data)
        except Exception:
            return {"raw_encrypted": data}

        for mode, kwargs in [
            (AES.MODE_ECB, {}),
            (AES.MODE_CBC, {"iv": b"\x00" * 16}),
        ]:
            try:
                cipher = AES.new(key, mode, **kwargs)
                dec = unpad(cipher.decrypt(raw), AES.block_size)
                return json.loads(dec.decode("utf-8"))
            except Exception:
                continue

        log.warning("Nao foi possivel decriptar a resposta AES -- retornando bruto")
        return {"raw_encrypted": data}

    def obter_usuario(self) -> dict:
        """
        Dados do usuario autenticado via /psec/validateToken (TreinoWeb, JSON em claro).
        O /adm/obter-usuario (API Gateway) retorna AES-encriptado com chave desconhecida;
        _decrypt_aes() esta disponivel para quando a chave for descoberta.
        """
        zw = PactoClient()
        return zw.obter_usuario()

    # -- FINANCEIRO -----------------------------------------------------------

    def _paginar(self, path: str, params: dict, max_pages: int = 50) -> list:
        """Pagina automaticamente um endpoint com content[] até esgotar."""
        todos, page = [], 0
        while page < max_pages:
            r = self._gw("GET", path, params={**params, "page": page, "size": 500})
            content = r.get("content", []) if isinstance(r, dict) else []
            if not content:
                break
            todos.extend(content)
            total_pages = r.get("totalPages")
            page += 1
            if total_pages is not None and page >= total_pages:
                break
        return todos

    # Mapeamento confirmado por varredura completa dos 3.918 lançamentos (2026-06-20):
    # 1=Despesa  3=Recebivel Debito  4=Transferencia  5=Estorno
    # 7=Conciliacao  8=Recebivel Cartao/Cheque  10=Retirada de lote
    # NOTA: tipo=2 (receita manual) não existe nesta academia —
    #       receitas de mensalidades vêm de /parcelas/{codPessoa}
    TIPOS_MOVCONTA = {
        1:  "Despesa",
        3:  "Recebivel Debito",
        4:  "Transferencia",
        5:  "Estorno",
        7:  "Conciliacao de Saldo",
        8:  "Recebivel Cartao/Cheque",
        10: "Retirada de Lote",
    }

    def movcontas(self, page: int = 0, size: int = 200) -> dict:
        """
        Lançamentos financeiros (contas a pagar e receber).
        Campos: codigo, descricao, pessoa, valor, dataLancamento,
                dataVencimento, dataCompetencia, tipoOperacao, nrParcela, conta.
        Ver TIPOS_MOVCONTA para o mapeamento completo de tipoOperacao.
        O filtro tipoOperacao via query param é ignorado pelo servidor — filtrar no cliente.
        """
        return self._gw("GET", "/v1/movconta", params={"page": page, "size": size})

    def receitas(self) -> list:
        """
        Recebíveis em movcontas: tipos 3 (débito) e 8 (cartão/cheque).
        Para receitas de mensalidades, use parcelas_cliente() com situacao=PG.
        """
        movs = self._paginar("/v1/movconta", {})
        return [m for m in movs if m.get("tipoOperacao") in (3, 8)]

    def despesas(self) -> list:
        """Lançamentos de despesa (tipoOperacao=1 — contas a pagar)."""
        movs = self._paginar("/v1/movconta", {})
        return [m for m in movs if m.get("tipoOperacao") == 1]

    def estornos(self) -> list:
        """Estornos registrados em movcontas (tipoOperacao=5)."""
        movs = self._paginar("/v1/movconta", {})
        return [m for m in movs if m.get("tipoOperacao") == 5]

    def planos(self, apenas_ativos: bool = False) -> list:
        """
        Planos de matrícula cadastrados.
        Campos: codigo, descricao, tipoPlano, vigenciaDe, vigenciaAte,
                ingressoAte, permitirAcessoSomenteNaEmpresaVende, ...
        """
        params = {"page": 0, "size": 500}
        if apenas_ativos:
            params["ativo"] = True
        return self._paginar("/planos", params)

    def contas(self) -> list:
        """Contas bancárias e caixas cadastrados."""
        return self._paginar("/v1/conta", {})

    def bancos(self) -> list:
        """Bancos cadastrados."""
        return self._paginar("/v1/banco", {})

    def centros_custo(self) -> list:
        """Centros de custo."""
        return self._paginar("/v1/centrocusto", {})

    def plano_contas(self) -> list:
        """Plano de contas (categorias financeiras)."""
        return self._paginar("/v1/planoconta", {})

    def meta_financeira(self) -> dict:
        """
        Meta financeira do período atual.
        Retorna faturamentoRecebido, faturamentoDevolvido, faturamentoLiquido,
        metas (lista por consultor/meta), produtosFaturamento.
        """
        r = self._gw("POST", "/v2-meta-financeira", json={})
        content = r.get("content", {}) if isinstance(r, dict) else {}
        jd = content.get("jsonDados", "{}")
        try:
            return json.loads(jd) if isinstance(jd, str) else jd
        except Exception:
            return content

    def resumo_financeiro(self) -> dict:
        """
        Resumo financeiro consolidado a partir de movcontas.
        Receitas aqui = recebíveis (tipos 3+8); despesas = tipo 1.
        Para receita de mensalidades, use parcelas_cliente() com situacao=PG.
        """
        log.info("Gerando resumo financeiro...")
        movs = self._paginar("/v1/movconta", {})

        receitas = [m for m in movs if m.get("tipoOperacao") in (3, 8)]
        despesas = [m for m in movs if m.get("tipoOperacao") == 1]
        estornos = [m for m in movs if m.get("tipoOperacao") == 5]

        total_receitas = sum(m.get("valor", 0) for m in receitas)
        total_despesas = sum(m.get("valor", 0) for m in despesas)
        total_estornos = sum(m.get("valor", 0) for m in estornos)

        por_tipo = {}
        for m in movs:
            t = m.get("tipoOperacao")
            por_tipo[t] = por_tipo.get(t, 0) + 1

        log.info(
            f"Financeiro: {len(movs)} lancamentos | "
            f"Recebiveis: R$ {total_receitas:,.2f} | Despesas: R$ {total_despesas:,.2f}"
        )

        return {
            "total_lancamentos": len(movs),
            "recebiveis_total":  round(total_receitas, 2),
            "despesas_total":    round(total_despesas, 2),
            "estornos_total":    round(total_estornos, 2),
            "saldo":             round(total_receitas - total_despesas, 2),
            "por_tipo":          por_tipo,
            "recebiveis":        receitas,
            "despesas":          despesas,
        }

    def grupo_risco(self) -> list:
        """
        Clientes em grupo de risco (potencial churn) gerado pelo BI.
        Campos: cliente (codigo), colaborador, codigo, empresa, nomeCliente.
        """
        r = self._gw("POST", "/v2-grupo-risco", json={})
        if not isinstance(r, dict) or r.get("erro"):
            return []
        content = r.get("content", {})
        jd = content.get("jsonDados", "{}")
        try:
            data = json.loads(jd) if isinstance(jd, str) else jd
            return data.get("listaClientes", [])
        except Exception:
            return []

    # -- CLIENTES ADM ---------------------------------------------------------

    def cliente(self, codigo_cliente: int) -> dict:
        """
        Retorna dados completos de um cliente pelo codigoCliente (ADM).
        Inclui: nome, CPF, telefones, endereço, email, foto, vínculos, avisos.
        O campo pessoa.codigo é o codPessoa usado em /parcelas/{codPessoa}.
        """
        r = self._gw("GET", f"/v1/cliente/{codigo_cliente}")
        return r.get("content", r) if isinstance(r, dict) else r

    def pessoa_de_cliente(self, codigo_cliente: int) -> int | None:
        """Retorna o codPessoa de um cliente (para uso em /parcelas/)."""
        c = self.cliente(codigo_cliente)
        return (c.get("pessoa") or {}).get("codigo")

    def parcelas_cliente(self, codigo_cliente: int, size: int = 100) -> list:
        """Parcelas de um cliente pelo codigoCliente (resolve codPessoa internamente)."""
        cod_pessoa = self.pessoa_de_cliente(codigo_cliente)
        if not cod_pessoa:
            return []
        r = self._gw("GET", f"/parcelas/{cod_pessoa}", params={"size": size})
        return r.get("content", []) if isinstance(r, dict) else []

    def faturamento_dia(self, d: date, max_clientes: int = 20000) -> dict:
        """
        Faturamento do dia somando parcelas pagas (situacao=PG) com dataPagamento = d.
        Itera /v1/cliente/{id} para obter codPessoa, depois /parcelas/{codPessoa}.
        max_clientes limita o range de codigoCliente varrido (padrão: 20000).
        """
        from datetime import timezone
        d_start = int(datetime(d.year, d.month, d.day, tzinfo=timezone.utc).timestamp() * 1000)
        d_end   = int(datetime(d.year, d.month, d.day + 1, tzinfo=timezone.utc).timestamp() * 1000)

        total = 0.0
        pagamentos = []
        log.info(f"Varrendo faturamento de {d}...")

        for cod in range(1, max_clientes + 1):
            cod_pessoa = self.pessoa_de_cliente(cod)
            if not cod_pessoa:
                continue
            parcelas = self._gw("GET", f"/parcelas/{cod_pessoa}", params={"size": 200})
            for p in (parcelas.get("content", []) if isinstance(parcelas, dict) else []):
                dt_pag = p.get("dataPagamento") or 0
                if d_start <= dt_pag < d_end and p.get("situacao") == "PG":
                    val = float(p.get("valor", 0) or 0)
                    total += val
                    pagamentos.append({
                        "codigoCliente": cod,
                        "codPessoa":     cod_pessoa,
                        "descricao":     p.get("descricao"),
                        "valor":         round(val, 2),
                        "dataPagamento": datetime.fromtimestamp(dt_pag / 1000,
                                         tz=timezone.utc).strftime("%Y-%m-%d %H:%M"),
                    })

        return {
            "data":         d.isoformat(),
            "total":        round(total, 2),
            "n_pagamentos": len(pagamentos),
            "pagamentos":   pagamentos,
        }

    def reconstruir_peso(self, cod_aluno: int,
                         dias_avencer_config: int = 10,
                         data_ref: date = None) -> dict:
        """
        Reconstrói o peso de risco de churn de um cliente usando a fórmula oficial do Pacto.

        Fórmula: peso = peso_vencimento + peso_faltas + peso_presenca  (range 3-8)

          peso_vencimento — proximidade do vencimento do contrato:
            2 se vence em <= dias_avencer_config dias (config empresa, padrão 10)
            1 caso contrário

          peso_faltas — faltas sequenciais (dias corridos sem acesso físico via catraca):
            1 se ausente 0-1 dias
            2 se ausente 2-3 dias
            3 se ausente 4+ dias
            (threshold configurável na empresa)

          peso_presenca — presença média nas últimas 4 semanas (janelas de 7 dias corridos):
            1 se >= 4 dias/semana
            2 se == 3 dias/semana
            3 se <= 2 dias/semana

        Fonte de acessos: consultarQuantidadeAcessosClientesAgrupadosDia (acesso físico / catraca).
        NÃO usa linha-tempo TreinoWeb (TREINOU = home workout, não conta).

        Args:
            cod_aluno: codigoCliente (campo 'cliente' em grupo_risco())
            dias_avencer_config: threshold "Ativo a Vencer" configurado na empresa (padrão 10)
            data_ref: data de referência (padrão: hoje)

        Returns:
            dict com peso_total, componentes individuais, vencimento, presença média e detalhes.
        """
        zw = PactoClient()
        ctx = zw.ctx

        hoje = data_ref or date.today()

        def ts(d: date) -> int:
            return int(datetime(d.year, d.month, d.day).timestamp() * 1000)

        def acessos_janela(ini: date, fim: date) -> int:
            r = zw._req("GET",
                f"/cliente/{ctx}/app/consultarQuantidadeAcessosClientesAgrupadosDia",
                params={"codigoAluno": cod_aluno, "username": "",
                        "dataInicial": ts(ini), "dataFinal": ts(fim)})
            if isinstance(r, dict) and "sucesso" in r:
                return sum(r["sucesso"].values())
            return 0

        # ── 1. peso_presenca: média nas últimas 4 semanas ──────────────
        total_4sem = acessos_janela(hoje - timedelta(weeks=4), hoje)
        media_semanal = total_4sem / 4.0
        if media_semanal >= 4:
            peso_presenca = 1
        elif media_semanal >= 3:
            peso_presenca = 2
        else:
            peso_presenca = 3

        # ── 2. peso_faltas: dias corridos sem acesso físico ────────────
        # Aproximação: verifica janelas curtas para determinar a faixa
        if acessos_janela(hoje - timedelta(days=2), hoje) > 0:
            peso_faltas = 1
            dias_ausente_aprox = "0-1"
        elif acessos_janela(hoje - timedelta(days=4), hoje) > 0:
            peso_faltas = 2
            dias_ausente_aprox = "2-3"
        else:
            peso_faltas = 3
            # estima há quantos dias não vai
            dias_ausente_aprox = "4+"
            for semanas in range(1, 27):
                ini = hoje - timedelta(weeks=semanas + 1)
                fim = hoje - timedelta(weeks=semanas)
                if acessos_janela(ini, fim) > 0:
                    dias_ausente_aprox = f"~{semanas * 7}d"
                    break

        # ── 3. peso_vencimento: próximo do vencimento? ─────────────────
        # Fonte primária: clienteSintetico.situacaocontrato.codigo == "AV" (ADM)
        # Fonte secundária: situacaoContratoZW == "A_VENCER" + contratoZW.vencimento (TreinoWeb)
        peso_vencimento = 1
        vencimento_str = "desconhecido"
        dias_para_vencer = None
        try:
            cli = self.cliente(cod_aluno)
            mat_zw = cli.get("matricula") if isinstance(cli, dict) else None

            # método 1: situacaocontrato no clienteSintetico (ADM)
            cs = (cli.get("clienteSintetico") or {}) if isinstance(cli, dict) else {}
            sit = (cs.get("situacaocontrato") or {})
            if (sit.get("codigo") == "AV") or (sit.get("descricao", "").lower() == "a vencer"):
                peso_vencimento = 2
                vencimento_str = "A Vencer (ADM)"

            # método 2: contratoZW.vencimento via TreinoWeb (mais preciso)
            if mat_zw:
                r_zw = zw._req("GET",
                    f"/psec/alunos/obter-aluno-completo-por-matricula/{mat_zw}")
                content = (r_zw.get("content", r_zw) if isinstance(r_zw, dict) else r_zw) or {}
                sit_zw = content.get("situacaoContratoZW", "")
                contrato_zw = content.get("contratoZW") or {}
                venc_ts = contrato_zw.get("vencimento")
                if venc_ts:
                    venc_date = datetime.fromtimestamp(venc_ts / 1000).date()
                    dias_para_vencer = (venc_date - hoje).days
                    vencimento_str = venc_date.strftime("%d/%m/%Y")
                    if 0 <= dias_para_vencer <= dias_avencer_config:
                        peso_vencimento = 2
                elif sit_zw in ("A_VENCER",):
                    peso_vencimento = 2
                    vencimento_str = "A Vencer (ZW)"
        except Exception as e:
            vencimento_str = f"erro: {e}"

        peso_total = peso_vencimento + peso_faltas + peso_presenca
        # garante range 3-8
        peso_total = max(3, min(8, peso_total))

        detalhes = (f"venc={peso_vencimento}({vencimento_str}"
                    f"{f', {dias_para_vencer}d' if dias_para_vencer is not None else ''})"
                    f" + faltas={peso_faltas}({dias_ausente_aprox})"
                    f" + pres={peso_presenca}({media_semanal:.1f}/sem)")

        return {
            "cod_aluno":           cod_aluno,
            "peso_calculado":      peso_total,
            "peso_vencimento":     peso_vencimento,
            "peso_faltas":         peso_faltas,
            "peso_presenca":       peso_presenca,
            "vencimento":          vencimento_str,
            "dias_para_vencer":    dias_para_vencer,
            "media_presenca_4sem": round(media_semanal, 2),
            "dias_ausente_aprox":  dias_ausente_aprox,
            "detalhes":            detalhes,
        }


# -- CRM SUPABASE (crm.territoriofit.com.br) ----------------------------------

# Namespace fixo para gerar UUIDs determinísticos a partir do código Pacto.
# Garante que o mesmo aluno sempre mapeia para o mesmo id no CRM, permitindo
# upsert seguro por primary key sem precisar de constraints extras.
_PACTO_NS = uuid.UUID("d7e8a9f0-cafe-5678-abcd-ef0123456789")


def _pacto_lead_id(pacto_codigo: int, tenant_id: str) -> str:
    return str(uuid.uuid5(_PACTO_NS, f"{tenant_id}:{pacto_codigo}"))


def _data_meta_iso(data_meta) -> str | None:
    """Converte dataMeta do meta-crm ('DD/MM/YYYY') para ISO 'YYYY-MM-DD'."""
    try:
        d, m, y = str(data_meta or "").strip().split("/")
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    except Exception:
        return None


class CRMClient:
    """
    Integração Pacto → CRM Território Fit (Supabase).
    Requer: pip install supabase
    """

    def __init__(self):
        try:
            from supabase import create_client
        except ImportError:
            raise RuntimeError(
                "Pacote 'supabase' não instalado. Execute: pip install supabase"
            )
        if not SUPABASE_KEY:
            raise RuntimeError("SUPABASE_KEY não definida no .env.integracao")
        self.sb = create_client(SUPABASE_URL, SUPABASE_KEY)
        self.tenant_id = CRM_TENANT_ID or self._bootstrap_tenant()

    # -- Setup -----------------------------------------------------------------

    def _bootstrap_tenant(self) -> str:
        """Cria ou recupera o tenant 'territorio-fit'. Salva o id no .env."""
        r = self.sb.table("tenants").select("id").eq("slug", "territorio-fit").execute()
        if r.data:
            tid = r.data[0]["id"]
        else:
            r = self.sb.table("tenants").insert({
                "name": "Território Fit",
                "slug": "territorio-fit",
                "is_active": True,
            }).execute()
            tid = r.data[0]["id"]
            log.info(f"Tenant criado: {tid}")

        # persiste no .env para não recriar nas próximas execuções
        env_path = Path(__file__).parent / ".env.integracao"
        content = env_path.read_text(encoding="utf-8")
        content = content.replace("CRM_TENANT_ID=", f"CRM_TENANT_ID={tid}")
        env_path.write_text(content, encoding="utf-8")
        log.info(f"Tenant: {tid}")
        return tid

    # -- Helpers ---------------------------------------------------------------

    @staticmethod
    def _tel(fones) -> str | None:
        if not fones:
            return None
        raw = fones[0].get("numero", "") if isinstance(fones[0], dict) else str(fones[0])
        digits = "".join(c for c in raw if c.isdigit())
        return digits or None

    @staticmethod
    def _email_addr(emails) -> str | None:
        if not emails:
            return None
        e = emails[0]
        return (e.get("email") if isinstance(e, dict) else str(e)) or None

    def _lead_row(self, pacto_codigo: int, aluno: dict, source: str = "pacto_aluno") -> dict:
        """Monta um dict pronto para upsert na tabela leads."""
        contrato = aluno.get("contratoZW") or {}
        venc_ts  = contrato.get("vencimento")
        venc_str = _ts_to_date(venc_ts).strftime("%Y-%m-%d") if venc_ts else None

        return {
            "id":        _pacto_lead_id(pacto_codigo, self.tenant_id),
            "name":      aluno.get("nome") or aluno.get("name"),
            "phone":     self._tel(aluno.get("fones") or []),
            "email":     self._email_addr(aluno.get("emails") or []),
            "source":    source,
            "status":    "cliente" if aluno.get("situacaoAluno") == "ATIVO" else "inativo",
            "tenant_id": self.tenant_id,
            "metadata":  {
                "pacto_codigo":    pacto_codigo,
                "pacto_matricula": aluno.get("matriculaZW") or aluno.get("id"),
                "pacto_plano":     (aluno.get("planoZW") or {}).get("descricao"),
                "pacto_situacao":  aluno.get("situacaoAluno"),
                "vencimento":      venc_str,
                "tipo_contrato":   contrato.get("tipo"),
                "synced_at":       datetime.now().isoformat(),
            },
        }

    def _upsert_batch(self, rows: list[dict]) -> int:
        """Upsert em lote de no máximo 200 linhas por chamada."""
        if not rows:
            return 0
        total = 0
        BATCH = 200
        for i in range(0, len(rows), BATCH):
            r = self.sb.table("leads").upsert(rows[i:i + BATCH]).execute()
            total += len(r.data or [])
        return total

    # -- Syncs -----------------------------------------------------------------

    def sync_alunos_ativos(self, pacto: "PactoClient") -> int:
        """Upsert de todos os alunos ativos do Pacto como leads/clientes no CRM."""
        log.info("CRM sync: alunos ativos...")
        alunos = pacto.alunos_ativos()
        rows = []
        for a in alunos:
            cod = a.get("codigoCliente")
            if cod:
                rows.append(self._lead_row(cod, a, source="pacto_aluno"))
        n = self._upsert_batch(rows)
        log.info(f"sync_alunos_ativos: {n}/{len(alunos)} ok")
        return n

    def sync_alunos_inativos(self, pacto: "PactoClient") -> int:
        """Upsert de todos os alunos inativos (ex-alunos) do Pacto como leads no CRM."""
        log.info("CRM sync: alunos inativos...")
        alunos = pacto.alunos_inativos()
        rows = []
        for a in alunos:
            cod = a.get("codigoCliente")
            if cod:
                rows.append(self._lead_row(cod, a, source="pacto_aluno"))
        n = self._upsert_batch(rows)
        log.info(f"sync_alunos_inativos: {n}/{len(alunos)} ok")
        return n

    def _stage_id_by_description(self, description: str, pipeline_id: str | None = None) -> str | None:
        """Busca o id da fase (sales_pipeline_stages) pelo slug (description)."""
        pipeline_id = pipeline_id or self._ensure_pipeline()
        r = self.sb.table("sales_pipeline_stages").select("id").eq(
            "pipeline_id", pipeline_id
        ).eq("description", description).limit(1).execute()
        return r.data[0]["id"] if r.data else None

    def sync_visitantes(self, adm: "PactoADMClient", data: date = None) -> int:
        """Upsert de visitantes 24h do Pacto como leads no CRM."""
        data = data or date.today()
        log.info(f"CRM sync: visitantes {data}...")
        stage_id = self._stage_id_by_description("VINTE_QUATRO_HORAS")
        r = adm.visitantes_24h(data)
        rows = []
        for v in r.get("visitantes", []):
            # chave = codigoCliente, o mesmo keyspace dos alunos e da base
            # completa de visitantes — visitante que matricular vira o mesmo lead
            cod = v.get("codigoCliente")
            if not cod:
                continue
            tel = "".join(c for c in (v.get("telefone") or "") if c.isdigit()) or None
            rows.append({
                "id":                _pacto_lead_id(cod, self.tenant_id),
                "name":              v.get("nome"),
                "phone":             tel,
                "email":             v.get("email"),
                "source":            "pacto_visitante",
                "status":            "lead",
                "sales_stage":       v.get("situacao"),
                "pipeline_stage_id": stage_id,
                "tenant_id":         self.tenant_id,
                "metadata": {
                    "pacto_codigo":     cod,
                    "pacto_matricula":  v.get("matricula"),
                    "pacto_situacao":   v.get("situacao"),
                    "pacto_consultora": v.get("consultora"),
                    "pacto_data_meta":  str(v.get("dataMeta", "")),
                    "data_visita":      _data_meta_iso(v.get("dataMeta")),
                    "synced_at":        datetime.now().isoformat(),
                },
            })
        n = self._upsert_batch(rows)
        log.info(f"sync_visitantes: {n}/{len(rows)} ok")
        return n

    def sync_grupo_risco(self, adm: "PactoADMClient") -> int:
        """Atualiza campo metadata.grupo_risco nos leads com dados de churn do Pacto."""
        log.info("CRM sync: grupo de risco...")
        clientes = adm.grupo_risco()
        if not clientes:
            log.info("Grupo de risco vazio.")
            return 0

        # busca leads em lotes para evitar URL too long (limite ~8KB)
        BATCH = 80
        existentes: dict = {}  # lid -> {metadata, name, status}
        pares = [(c["cliente"], _pacto_lead_id(c["cliente"], self.tenant_id))
                 for c in clientes if c.get("cliente")]
        ids = [lid for _, lid in pares]
        for i in range(0, len(ids), BATCH):
            r = self.sb.table("leads").select("id,name,status,metadata").in_(
                "id", ids[i:i+BATCH]
            ).execute()
            for row in (r.data or []):
                existentes[row["id"]] = row

        rows = []
        for c in clientes:
            cod = c.get("cliente")
            if not cod:
                continue
            lid = _pacto_lead_id(cod, self.tenant_id)
            if lid not in existentes:
                continue  # lead ainda não sincronizado; sync_alunos_ativos o criará
            lead = existentes[lid]
            meta = dict(lead.get("metadata") or {})
            meta["grupo_risco"] = {
                "peso":        c.get("peso"),
                "colaborador": c.get("colaborador"),
                "synced_at":   datetime.now().isoformat(),
            }
            rows.append({
                "id":        lid,
                "name":      lead.get("name"),
                "status":    lead.get("status"),
                "metadata":  meta,
                "tenant_id": self.tenant_id,
            })

        n = self._upsert_batch(rows)
        log.info(f"sync_grupo_risco: {n}/{len(clientes)} ok ({len(clientes)-n} nao encontrados no CRM)")
        return n

    def sync_ultimo_acesso(self, pacto: "PactoClient") -> int:
        """
        Atualiza metadata.ultimo_acesso (data do ultimo acesso fisico na catraca)
        de todos os alunos ativos. Fonte: linha-tempo por aluno (evento 'Chegou'),
        1 request por aluno (~0.4s cada; ~15min na base de ~2000 ativos).
        Quem nunca acessou fica com ultimo_acesso=null (frontend mostra 'sem acesso').
        """
        log.info("CRM sync: ultimo acesso catraca...")
        alunos = pacto.alunos_ativos()

        # busca metadata atual dos leads em lotes (merge, padrao sync_grupo_risco)
        BATCH = 80
        existentes: dict = {}
        pares = [(a, _pacto_lead_id(a["codigoCliente"], self.tenant_id))
                 for a in alunos if a.get("codigoCliente")]
        ids = [lid for _, lid in pares]
        for i in range(0, len(ids), BATCH):
            r = self.sb.table("leads").select("id,name,status,metadata").in_(
                "id", ids[i:i+BATCH]
            ).execute()
            for row in (r.data or []):
                existentes[row["id"]] = row

        rows, erros = [], 0
        for a, lid in pares:
            if lid not in existentes:
                continue  # lead ainda nao sincronizado; sync_alunos_ativos o criara
            mat = a.get("matriculaZW") or a.get("id")
            if not mat:
                continue
            ultimo, falhou = None, False
            for tentativa in (1, 2):
                try:
                    ultimo = pacto.ultimo_acesso_catraca(mat)
                    falhou = False
                    break
                except Exception as e:
                    falhou = True
                    if tentativa == 2:
                        erros += 1
                        log.warning(f"ultimo_acesso mat={mat}: {e}")
            if falhou:
                continue  # nao sobrescreve dado anterior com falha de rede
            lead = existentes[lid]
            meta = dict(lead.get("metadata") or {})
            meta["ultimo_acesso"] = ultimo.strftime("%Y-%m-%d") if ultimo else None
            meta["ultimo_acesso_synced_at"] = datetime.now().isoformat()
            rows.append({
                "id":        lid,
                "name":      lead.get("name"),
                "status":    lead.get("status"),
                "metadata":  meta,
                "tenant_id": self.tenant_id,
            })

        n = self._upsert_batch(rows)
        log.info(f"sync_ultimo_acesso: {n}/{len(pares)} ok ({erros} falhas de consulta)")
        return n

    def sync_aniversariantes(self, pacto: "PactoClient") -> int:
        """
        Marca metadata.aniversariante_mes_atual (+ data_nascimento) nos alunos
        ATIVOS e INATIVOS que fazem aniversario no mes corrente e desmarca quem
        saiu do mes. Fonte: dataNascimento da listagem /psec/alunos.
        Para aniversariante sem ultimo_acesso no metadata (inativos ficam fora
        do sync_ultimo_acesso) busca tambem o ultimo acesso na catraca.
        Alimenta a coluna Aniversariantes do Kanban de Alunos.
        """
        log.info("CRM sync: aniversariantes do mes (ativos + inativos)...")
        alunos = pacto.todos_alunos()
        mes = date.today().month

        aniversariantes = []
        for a in alunos:
            if a.get("situacaoAluno") not in ("ATIVO", "INATIVO"):
                continue
            cod = a.get("codigoCliente")
            nasc = _ts_to_date(a.get("dataNascimento")) if a.get("dataNascimento") else None
            if cod and nasc and nasc.month == mes:
                aniversariantes.append((a, nasc, _pacto_lead_id(cod, self.tenant_id)))
        ids_novos = {lid for _, _, lid in aniversariantes}

        rows = []

        # desmarca quem esta com a flag mas nao faz aniversario neste mes
        r = self.sb.table("leads").select("id,name,status,metadata").eq(
            "source", "pacto_aluno"
        ).eq("metadata->>aniversariante_mes_atual", "true").limit(5000).execute()
        for row in (r.data or []):
            if row["id"] in ids_novos:
                continue
            meta = dict(row.get("metadata") or {})
            meta["aniversariante_mes_atual"] = False
            rows.append({
                "id":        row["id"],
                "name":      row.get("name"),
                "status":    row.get("status"),
                "metadata":  meta,
                "tenant_id": self.tenant_id,
            })

        # metadata atual dos aniversariantes em lotes (merge, padrao sync_grupo_risco)
        BATCH = 80
        existentes: dict = {}
        ids = [lid for _, _, lid in aniversariantes]
        for i in range(0, len(ids), BATCH):
            r = self.sb.table("leads").select("id,name,status,metadata").in_(
                "id", ids[i:i+BATCH]
            ).execute()
            for row in (r.data or []):
                existentes[row["id"]] = row

        buscou_acesso = 0
        for a, nasc, lid in aniversariantes:
            lead = existentes.get(lid)
            if lead is None:
                # ex-aluno ainda nao sincronizado como lead: cria a linha completa
                row = self._lead_row(a["codigoCliente"], a)
                meta = row["metadata"]
            else:
                meta = dict(lead.get("metadata") or {})
                row = {
                    "id":        lid,
                    "name":      lead.get("name"),
                    "status":    lead.get("status"),
                    "metadata":  meta,
                    "tenant_id": self.tenant_id,
                }
            meta["aniversariante_mes_atual"] = True
            meta["data_nascimento"] = nasc.strftime("%Y-%m-%d")
            # ultimo acesso na catraca p/ quem nunca recebeu (1 request por aluno)
            if "ultimo_acesso_synced_at" not in meta:
                mat = a.get("matriculaZW") or a.get("id")
                if mat:
                    try:
                        ua = pacto.ultimo_acesso_catraca(mat)
                        meta["ultimo_acesso"] = ua.strftime("%Y-%m-%d") if ua else None
                        meta["ultimo_acesso_synced_at"] = datetime.now().isoformat()
                        buscou_acesso += 1
                    except Exception as e:
                        log.warning(f"ultimo_acesso aniversariante mat={mat}: {e}")
            rows.append(row)

        n = self._upsert_batch(rows)
        log.info(f"sync_aniversariantes: {len(aniversariantes)} aniversariantes no mes "
                 f"(ativos+inativos), {n} leads atualizados, {buscou_acesso} acessos buscados")
        return n

    def sync_contratos_a_vencer(self, pacto: "PactoClient", dias: int = 30) -> int:
        """Upsert de clientes com contrato a vencer, marcando alerta no metadata."""
        log.info(f"CRM sync: contratos a vencer ({dias}d)...")
        a_vencer = pacto.contratos_a_vencer(dias=dias)
        rows = []
        for a in a_vencer:
            cod = a.get("codigoCliente")
            if not cod:
                continue
            row = self._lead_row(cod, a, source="pacto_aluno")
            row["metadata"]["alerta"]          = "contrato_a_vencer"
            row["metadata"]["dias_restantes"]  = a.get("dias_restantes")
            rows.append(row)
        n = self._upsert_batch(rows)
        log.info(f"sync_contratos_a_vencer: {n}/{len(a_vencer)} ok")
        return n

    @staticmethod
    def _month_bounds(ref: date | None = None) -> tuple[datetime, datetime]:
        """Retorna (inicio, fim) do mes calendario de `ref` (padrao: hoje)."""
        ref = ref or date.today()
        inicio = datetime(ref.year, ref.month, 1)
        if ref.month == 12:
            fim = datetime(ref.year + 1, 1, 1) - timedelta(seconds=1)
        else:
            fim = datetime(ref.year, ref.month + 1, 1) - timedelta(seconds=1)
        return inicio, fim

    def sync_renovacao_mes_atual(self, pacto: "PactoClient") -> int:
        """
        Marca na etapa 'Renovacao' os alunos ATIVOS cujo contrato vence dentro
        do mes calendario atual. Quem sai da janela do mes e removido da etapa.
        """
        log.info("CRM sync: renovacao (vencimento no mes atual)...")
        stage_id = self._stage_id_by_description("RENOVACAO")
        if not stage_id:
            log.warning("Etapa RENOVACAO nao encontrada no pipeline.")
            return 0

        inicio, fim = self._month_bounds()
        alunos = pacto.alunos_ativos()
        rows, ids_no_mes = [], []
        for a in alunos:
            cod = a.get("codigoCliente")
            if not cod:
                continue
            ts   = (a.get("contratoZW") or {}).get("vencimento")
            venc = _ts_to_date(ts)
            if venc and inicio <= venc <= fim:
                row = self._lead_row(cod, a, source="pacto_aluno")
                row["pipeline_stage_id"] = stage_id
                row["metadata"]["vencimento_mes_atual"] = venc.strftime("%Y-%m-%d")
                rows.append(row)
                ids_no_mes.append(row["id"])

        n = self._upsert_batch(rows)

        # Tira da etapa quem nao vence mais neste mes (evita ficar preso pra sempre)
        query = self.sb.table("leads").update({"pipeline_stage_id": None}).eq(
            "pipeline_stage_id", stage_id
        ).eq("source", "pacto_aluno")
        if ids_no_mes:
            query = query.not_.in_("id", ids_no_mes)
        query.execute()

        log.info(f"sync_renovacao_mes_atual: {n}/{len(rows)} vencendo em {inicio:%m/%Y}")
        return n

    def sync_renovacoes_status(self) -> dict:
        """
        Detecta renovacoes lancadas direto no Pacto e marca 'renovado' na
        pagina de Renovacao do CRM (tabela renovacoes).

        Logica: sync_alunos_ativos ja gravou hoje o vencimento do contrato
        ATUAL de cada aluno em leads.metadata.vencimento. Se esse vencimento
        ficou mais de 15 dias DEPOIS da previsao registrada na renovacao, um
        contrato novo foi lancado no Pacto -> renovou. A margem de 15 dias
        evita falso positivo por diferenca de poucos dias entre a previsao
        (vinda da planilha) e o vencimento real no Pacto.

        So promove status (pendente/negociando/aguardando_pagamento ->
        renovado); nunca rebaixa nem toca em nao_renova (controle manual).
        """
        log.info("CRM sync: renovacoes (deteccao de renovacao no Pacto)...")
        MARGEM_DIAS = 15

        r = self.sb.table("renovacoes").select(
            "id,tenant_id,nome,previsao,lead_id,mes_referencia,status"
        ).not_.in_("status", ["renovado", "nao_renova"]).execute()
        abertas = [x for x in (r.data or []) if x.get("lead_id") and x.get("previsao")]
        if not abertas:
            log.info("sync_renovacoes_status: nenhuma renovacao aberta com lead")
            return {"abertas": 0, "renovadas": 0}

        # vencimento atual dos leads (escrito hoje pelo sync_alunos_ativos)
        venc_por_lead: dict[str, str] = {}
        lead_ids = list({x["lead_id"] for x in abertas})
        for i in range(0, len(lead_ids), 200):
            rl = self.sb.table("leads").select("id,metadata").in_(
                "id", lead_ids[i:i + 200]
            ).execute()
            for ld in rl.data or []:
                v = (ld.get("metadata") or {}).get("vencimento")
                if v:
                    venc_por_lead[ld["id"]] = v

        renovadas = 0
        for ren in abertas:
            venc_atual = venc_por_lead.get(ren["lead_id"])
            if not venc_atual:
                continue
            try:
                d_venc = date.fromisoformat(venc_atual)
                d_prev = date.fromisoformat(ren["previsao"])
            except ValueError:
                continue
            if (d_venc - d_prev).days <= MARGEM_DIAS:
                continue
            # contrato novo lancado no Pacto -> marca renovado + timeline
            self.sb.table("renovacoes").update(
                {"status": "renovado"}
            ).eq("id", ren["id"]).execute()
            self.sb.table("renovacao_eventos").insert({
                "tenant_id":      ren["tenant_id"],
                "renovacao_id":   ren["id"],
                "tipo":           "status",
                "descricao":      (f"Renovação detectada no Pacto "
                                   f"(novo vencimento: {d_venc:%d/%m/%Y})"),
                "registrado_por": "Sync Pacto",
            }).execute()
            renovadas += 1
            log.info(f"  renovado: {ren['nome']} ({ren['mes_referencia']}) "
                     f"venc {ren['previsao']} -> {venc_atual}")

        log.info(f"sync_renovacoes_status: {renovadas} renovacoes detectadas "
                 f"de {len(abertas)} abertas")
        return {"abertas": len(abertas), "renovadas": renovadas}

    def sync_agendamentos_status(self) -> dict:
        """
        Detecta matriculas lancadas no Pacto e marca 'fechou' na pagina de
        Agendamentos do CRM (tabela agendamentos).

        Duas etapas:
          1. agendamentos sem lead_id ganham vinculo pelo final do telefone
             (8 digitos) — o lead pode ter sido criado depois do agendamento.
             Ambiguo (2+ leads com o mesmo final) fica sem vinculo.
          2. agendamento aberto (fechou nulo/false) cujo lead virou
             cliente/inadimplente -> fechou=true + evento na timeline.

        So promove; nunca rebaixa (fechou=false continua controle manual).
        """
        log.info("CRM sync: agendamentos (deteccao de fechamento no Pacto)...")
        r = self.sb.table("agendamentos").select(
            "id,tenant_id,nome,telefone,lead_id,mes_referencia,fechou"
        ).not_.is_("fechou", "true").execute()
        abertos = r.data or []
        if not abertos:
            log.info("sync_agendamentos_status: nenhum agendamento aberto")
            return {"abertos": 0, "vinculados": 0, "fechados": 0}

        # 1. vincula lead pelo final do telefone
        vinculados = 0
        for ag in abertos:
            tel = ag.get("telefone") or ""
            if ag.get("lead_id") or len(tel) < 8:
                continue
            rl = self.sb.table("leads").select("id").like(
                "phone", f"%{tel[-8:]}").limit(2).execute()
            if rl.data and len(rl.data) == 1:
                ag["lead_id"] = rl.data[0]["id"]
                self.sb.table("agendamentos").update(
                    {"lead_id": ag["lead_id"]}).eq("id", ag["id"]).execute()
                vinculados += 1

        # 2. lead ativo no Pacto = fechou
        status_por_lead: dict[str, str] = {}
        lead_ids = list({a["lead_id"] for a in abertos if a.get("lead_id")})
        for i in range(0, len(lead_ids), 200):
            rl = self.sb.table("leads").select("id,status").in_(
                "id", lead_ids[i:i + 200]).execute()
            for ld in rl.data or []:
                status_por_lead[ld["id"]] = ld.get("status") or ""

        fechados = 0
        for ag in abertos:
            if status_por_lead.get(ag.get("lead_id") or "") not in ("cliente", "inadimplente"):
                continue
            self.sb.table("agendamentos").update(
                {"fechou": True}).eq("id", ag["id"]).execute()
            self.sb.table("agendamento_eventos").insert({
                "tenant_id":      ag["tenant_id"],
                "agendamento_id": ag["id"],
                "tipo":           "auto",
                "descricao":      "Fechamento detectado no Pacto (aluno ativo)",
                "registrado_por": "Sync Pacto",
            }).execute()
            fechados += 1
            log.info(f"  fechou: {ag['nome']} ({ag['mes_referencia']})")

        log.info(f"sync_agendamentos_status: {fechados} fechamentos, "
                 f"{vinculados} vinculos novos, {len(abertos)} abertos")
        return {"abertos": len(abertos), "vinculados": vinculados, "fechados": fechados}

    def sync_leads_acompanhamento_status(self) -> dict:
        """
        Pagina Leads (acompanhamento): vincula linhas sem lead_id pelo final do
        telefone e marca fechou=true quando o lead virou cliente/inadimplente
        no Pacto. Mesma logica do sync_agendamentos_status; so promove.
        """
        log.info("CRM sync: leads acompanhamento (fechamento no Pacto)...")
        r = self.sb.table("leads_acompanhamento").select(
            "id,nome,telefone,lead_id,mes_referencia,fechou"
        ).not_.is_("fechou", "true").execute()
        abertos = r.data or []
        if not abertos:
            return {"abertos": 0, "vinculados": 0, "fechados": 0}

        vinculados = 0
        for ln in abertos:
            tel = ln.get("telefone") or ""
            if ln.get("lead_id") or len(tel) < 8:
                continue
            rl = self.sb.table("leads").select("id").like(
                "phone", f"%{tel[-8:]}").limit(2).execute()
            if rl.data and len(rl.data) == 1:
                ln["lead_id"] = rl.data[0]["id"]
                self.sb.table("leads_acompanhamento").update(
                    {"lead_id": ln["lead_id"]}).eq("id", ln["id"]).execute()
                vinculados += 1

        status_por_lead: dict[str, str] = {}
        lead_ids = list({x["lead_id"] for x in abertos if x.get("lead_id")})
        for i in range(0, len(lead_ids), 200):
            rl = self.sb.table("leads").select("id,status").in_(
                "id", lead_ids[i:i + 200]).execute()
            for ld in rl.data or []:
                status_por_lead[ld["id"]] = ld.get("status") or ""

        fechados = 0
        for ln in abertos:
            if status_por_lead.get(ln.get("lead_id") or "") not in ("cliente", "inadimplente"):
                continue
            self.sb.table("leads_acompanhamento").update(
                {"fechou": True}).eq("id", ln["id"]).execute()
            fechados += 1
            log.info(f"  fechou: {ln.get('nome') or ln.get('telefone')} ({ln['mes_referencia']})")

        log.info(f"sync_leads_acompanhamento_status: {fechados} fechados, "
                 f"{vinculados} vinculos novos, {len(abertos)} abertos")
        return {"abertos": len(abertos), "vinculados": vinculados, "fechados": fechados}

    def sync_consultora_vinculo(self, adm: "PactoADMClient") -> int:
        """
        Grava leads.metadata.consultora com a consultora do VINCULO do cliente
        no Pacto (GET /v1/cliente/{codigo} -> vinculos[tipoVinculo=CONSULTOR]).

        Escopo: inadimplentes (status=inadimplente + quem tem linha em
        parcelas_atrasadas) — alimenta o resumo por consultora do kanban de
        Inadimplencia. 1 request por cliente (~100-150/dia).

        ATENCAO: o upsert do sync_alunos_ativos SUBSTITUI o metadata inteiro,
        entao este sync precisa rodar DEPOIS dele no diario (padrao
        ultimo_acesso/aniversariantes).
        """
        log.info("CRM sync: consultora do vinculo (inadimplentes)...")
        alvo: dict[str, dict] = {}
        r = self.sb.table("leads").select("id,metadata").eq("status", "inadimplente").execute()
        for ld in r.data or []:
            alvo[ld["id"]] = ld
        rp = self.sb.table("parcelas_atrasadas").select("lead_id").not_.is_(
            "lead_id", "null").execute()
        extras = list({x["lead_id"] for x in (rp.data or [])} - set(alvo))
        for i in range(0, len(extras), 100):
            rl = self.sb.table("leads").select("id,metadata").in_("id", extras[i:i + 100]).execute()
            for ld in rl.data or []:
                alvo[ld["id"]] = ld

        # nomes canonicos iguais aos usados em renovacoes/agendamentos
        CANON = {"kelytta": "Kellyta", "kellyta": "Kellyta",
                 "nathalia": "Nathalia", "nathy": "Nathalia",
                 "raiane": "Raiane", "rai": "Raiane",
                 "ly": "Lyandra", "lyandra": "Lyandra"}
        n = falhas = 0
        for ld in alvo.values():
            meta = ld.get("metadata") or {}
            cod = meta.get("pacto_codigo")
            if not cod:
                continue
            try:
                r = adm._gw("GET", f"/v1/cliente/{cod}", timeout=30)
                vincs = (r.get("content") or {}).get("vinculos") or []
            except Exception as e:
                log.warning(f"vinculo cliente {cod}: {e}")
                falhas += 1
                continue
            nome = next((((v.get("colaborador") or {}).get("nome") or "")
                         for v in vincs if v.get("tipoVinculo") == "CONSULTOR"), "")
            if not nome.strip():
                continue
            primeiro = nome.strip().split()[0]
            consultora = CANON.get(primeiro.lower(), primeiro.title())
            if meta.get("consultora") == consultora:
                continue
            meta["consultora"] = consultora
            self.sb.table("leads").update({"metadata": meta}).eq("id", ld["id"]).execute()
            n += 1
        log.info(f"sync_consultora_vinculo: {n} atualizados de {len(alvo)} "
                 f"({falhas} falhas)")
        return n

    def sync_inadimplentes(self, pacto: "PactoClient", adm: "PactoADMClient") -> int:
        """
        Upsert de inadimplentes (contrato vencido) com alerta no metadata +
        dados da parcela atrasada mais antiga (codigo, vencimento, numero de
        tentativas de cobranca), buscados via /parcelas/{codigoPessoa}.

        NOTA (2026-07-02): contrato vencido != parcela em atraso no Pacto.
        Quem tem parcela realmente atrasada (situacao EA + vencimento no
        passado) e coberto por sync_parcelas_atrasadas(), que escaneia o
        grupo de risco (peso >= 5) + inadimplentes aluno a aluno.
        """
        log.info("CRM sync: inadimplentes...")
        inadimpl = pacto.inadimplentes()
        rows = []
        for a in inadimpl:
            cod = a.get("codigoCliente")
            if not cod:
                continue
            row = self._lead_row(cod, a, source="pacto_aluno")
            row["status"]                    = "inadimplente"
            row["metadata"]["alerta"]        = "inadimplente"
            row["metadata"]["dias_atraso"]   = a.get("dias_atraso")

            # Parcela em aberto e vencida (situacao=EA "Em aberto" + dataVencimento no
            # passado) mais antiga: codigo, vencimento, tentativas de cobranca.
            # IMPORTANTE:
            #  - O codigo correto e "EA", nao "AB" (confirmado via situacaoDescricao
            #    == "Em aberto" em dados reais).
            #  - Usa codigoPessoa do proprio TreinoWeb -- pessoa_de_cliente()/ADM
            #    retorna sempre codigo=1 (registro generico da empresa, nao da
            #    pessoa) neste token, entao NAO usar adm.parcelas_cliente()/
            #    pessoa_de_cliente() aqui.
            #  - "EA" sozinho inclui parcelas futuras do plano de pagamento (ainda
            #    nao venceram) -- filtra por dataVencimento < hoje pra pegar só
            #    quem esta de fato atrasado.
            cod_pessoa = a.get("codigoPessoa")
            try:
                if not cod_pessoa:
                    raise ValueError("sem codigoPessoa")
                hoje_ms = int(datetime.now().timestamp() * 1000)
                parcelas = adm._gw("GET", f"/parcelas/{cod_pessoa}", params={"size": 100})
                parcelas = parcelas.get("content", []) if isinstance(parcelas, dict) else []
                atrasadas = [
                    p for p in parcelas
                    if p.get("situacao") == "EA" and (p.get("dataVencimento") or 0) < hoje_ms
                ]
                atrasadas.sort(key=lambda p: p.get("dataVencimento") or 0)
                if atrasadas:
                    p = atrasadas[0]
                    venc_ts = p.get("dataVencimento")
                    venc_date = _ts_to_date(venc_ts) if venc_ts else None
                    row["metadata"]["parcela_atrasada"] = {
                        "codigo":               p.get("codigo"),
                        "vencimento":           venc_date.strftime("%Y-%m-%d") if venc_date else None,
                        "nr_tentativas":        p.get("nrTentativas"),
                        "valor":                p.get("valor"),
                        "total_parcelas_abertas": len(atrasadas),
                    }
            except Exception as e:
                log.warning(f"parcelas em aberto falharam pro cliente {cod}: {e}")

            rows.append(row)
        n = self._upsert_batch(rows)
        log.info(f"sync_inadimplentes: {n}/{len(inadimpl)} ok")
        return n

    # -- Parcelas atrasadas (situacao EA + vencimento no passado) -------------

    _PARCELA_NS = uuid.UUID("c0b4a5d6-cafe-5678-abcd-9876543210ab")

    @staticmethod
    def _motivo_falha_parcela(p: dict) -> str:
        """
        Motivo (derivado) pelo qual a parcela nao foi cobrada.

        A API do Pacto NAO expoe o retorno da operadora (investigado em
        2026-07-02): o objeto parcela so traz nrTentativas; os BIs
        /v2-pendencias e /v2-inadimplencia retornam 500 neste token; e o
        gateway nao roteia nenhum endpoint de historico de cobranca
        (/cobranca/*, /transacoes, /parcelas/{id}/tentativas -- todos 404).
        Derivamos o melhor motivo possivel do que existe na parcela.
        """
        nt = p.get("nrTentativas") or 0
        desc = (p.get("descricao") or "").upper()
        if nt > 0:
            plural = "s" if nt != 1 else ""
            return f"Cobranca automatica sem sucesso ({nt} tentativa{plural})"
        if "RENEGOCIA" in desc:
            return "Parcela renegociada nao paga"
        return "Sem tentativa de cobranca registrada (boleto/pix/manual nao pago)"

    def scan_parcelas_atrasadas(self, pacto: "PactoClient", adm: "PactoADMClient",
                                peso_minimo: int = 5,
                                todos_ativos: bool = False) -> list[dict]:
        """
        Escaneia aluno a aluno as parcelas EA ("Em aberto") com dataVencimento
        no passado -- as que o sistema tentou cobrar e nao conseguiu.

        Base do scan (nao existe endpoint em lote; /v1/movconta ignora filtros
        server-side, entao o filtro e todo client-side):
          - padrao: grupo de risco com peso >= peso_minimo (~724 com peso>=5)
            + inadimplentes de contrato vencido (~46) -- uniao ~740 alunos.
          - todos_ativos=True: escaneia a base ativa inteira (~2000 alunos,
            ~2x mais lento; preparado pra rodar de madrugada).

        Retorna [{"aluno": <aluno TreinoWeb>, "parcelas": [parcela EA vencida...]}]
        com as parcelas ordenadas da mais antiga pra mais recente.
        """
        ativos  = pacto.alunos_ativos()
        por_cod = {a.get("codigoCliente"): a for a in ativos if a.get("codigoCliente")}
        hoje    = datetime.now()
        hoje_ms = int(hoje.timestamp() * 1000)

        # inadimplentes (contrato vencido) calculados localmente pra nao
        # refazer o fetch de alunos_ativos()
        inad_cods = set()
        for cod, a in por_cod.items():
            venc = _ts_to_date((a.get("contratoZW") or {}).get("vencimento"))
            if venc and venc < hoje:
                inad_cods.add(cod)

        if todos_ativos:
            alvo = set(por_cod)
        else:
            risco = adm.grupo_risco()
            alvo  = {c["cliente"] for c in risco
                     if c.get("cliente") and (c.get("peso") or 0) >= peso_minimo}
            alvo |= inad_cods
            alvo &= set(por_cod)

        log.info(f"scan_parcelas_atrasadas: escaneando {len(alvo)} alunos "
                 f"({'todos ativos' if todos_ativos else f'risco peso>={peso_minimo} + inadimplentes'})...")

        achados = []
        for i, cod in enumerate(sorted(alvo)):
            aluno      = por_cod[cod]
            cod_pessoa = aluno.get("codigoPessoa")
            if not cod_pessoa:
                continue
            # 1 retry: na base completa ~3/2000 clientes dao read timeout (20s)
            for tentativa in (1, 2):
                try:
                    r = adm._gw("GET", f"/parcelas/{cod_pessoa}", params={"size": 100})
                    parcelas = r.get("content", []) if isinstance(r, dict) else []
                    atrasadas = [
                        p for p in parcelas
                        if p.get("situacao") == "EA" and (p.get("dataVencimento") or 0) < hoje_ms
                    ]
                    if atrasadas:
                        atrasadas.sort(key=lambda p: p.get("dataVencimento") or 0)
                        achados.append({"aluno": aluno, "parcelas": atrasadas})
                    break
                except Exception as e:
                    if tentativa == 2:
                        log.warning(f"scan parcelas: erro no cliente {cod}: {e}")
            if (i + 1) % 100 == 0:
                log.info(f"scan parcelas: {i + 1}/{len(alvo)} alunos, "
                         f"{len(achados)} com parcela atrasada")

        total_parcelas = sum(len(x["parcelas"]) for x in achados)
        log.info(f"scan_parcelas_atrasadas: {len(achados)} alunos com "
                 f"{total_parcelas} parcelas atrasadas (de {len(alvo)} escaneados)")
        return achados

    def sync_parcelas_atrasadas(self, pacto: "PactoClient", adm: "PactoADMClient",
                                peso_minimo: int = 5,
                                todos_ativos: bool = False) -> dict:
        """
        Scan + gravacao das parcelas atrasadas no Supabase:
          1. tabela parcelas_atrasadas: 1 linha por parcela (nome, matricula,
             codigo da parcela, vencimento, motivo da falha, tentativas);
             parcelas que sairam do atraso (pagas/renegociadas) sao removidas.
          2. leads: metadata.parcela_atrasada + alerta/status inadimplente
             (merge com metadata existente pra nao perder grupo_risco etc);
             quem nao tem mais parcela atrasada tem a flag limpa.
        """
        achados    = self.scan_parcelas_atrasadas(pacto, adm, peso_minimo, todos_ativos)
        agora      = datetime.now()
        scan_start = agora.isoformat()

        # -- 1) tabela parcelas_atrasadas -----------------------------------
        rows_parc = []
        por_lead  = {}  # lead_id -> (aluno, parcelas)
        for item in achados:
            a    = item["aluno"]
            cod  = a.get("codigoCliente")
            lid  = _pacto_lead_id(cod, self.tenant_id)
            por_lead[lid] = item
            for p in item["parcelas"]:
                venc = _ts_to_date(p.get("dataVencimento"))
                rows_parc.append({
                    "id":             str(uuid.uuid5(self._PARCELA_NS, f"{self.tenant_id}:{p.get('codigo')}")),
                    "tenant_id":      self.tenant_id,
                    "lead_id":        lid,
                    "nome_aluno":     a.get("nome"),
                    "matricula":      str(a.get("matriculaZW") or a.get("id") or ""),
                    "codigo_cliente": cod,
                    "codigo_pessoa":  a.get("codigoPessoa"),
                    "parcela_codigo": p.get("codigo"),
                    "contrato":       p.get("contrato"),
                    "descricao":      p.get("descricao"),
                    "valor":          p.get("valor"),
                    "data_vencimento": venc.strftime("%Y-%m-%d") if venc else None,
                    "dias_atraso":    (agora - venc).days if venc else None,
                    "nr_tentativas":  p.get("nrTentativas") or 0,
                    "motivo_falha":   self._motivo_falha_parcela(p),
                    "synced_at":      scan_start,
                })
        for i in range(0, len(rows_parc), 200):
            self.sb.table("parcelas_atrasadas").upsert(rows_parc[i:i + 200]).execute()
        # remove parcelas que nao apareceram neste scan (pagas/renegociadas)
        self.sb.table("parcelas_atrasadas").delete().eq(
            "tenant_id", self.tenant_id
        ).lt("synced_at", scan_start).execute()

        # -- 2) leads: flag + resumo no metadata (merge, padrao sync_grupo_risco)
        BATCH = 80
        ids = list(por_lead.keys())
        existentes: dict = {}
        for i in range(0, len(ids), BATCH):
            r = self.sb.table("leads").select("id,name,status,metadata").in_(
                "id", ids[i:i + BATCH]
            ).execute()
            for row in (r.data or []):
                existentes[row["id"]] = row

        rows_leads = []
        for lid, item in por_lead.items():
            if lid not in existentes:
                continue  # lead ainda nao sincronizado; sync_alunos_ativos o criara
            a, parcelas = item["aluno"], item["parcelas"]
            p    = parcelas[0]  # mais antiga
            venc = _ts_to_date(p.get("dataVencimento"))
            lead = existentes[lid]
            meta = dict(lead.get("metadata") or {})
            meta["alerta"] = "inadimplente"
            meta["parcela_atrasada"] = {
                "codigo":                 p.get("codigo"),
                "vencimento":             venc.strftime("%Y-%m-%d") if venc else None,
                "dias_atraso":            (agora - venc).days if venc else None,
                "nr_tentativas":          p.get("nrTentativas") or 0,
                "valor":                  p.get("valor"),
                "motivo_falha":           self._motivo_falha_parcela(p),
                "matricula":              str(a.get("matriculaZW") or a.get("id") or ""),
                "total_parcelas_abertas": len(parcelas),
            }
            rows_leads.append({
                "id":        lid,
                "name":      lead.get("name"),
                "status":    "inadimplente",
                "metadata":  meta,
                "tenant_id": self.tenant_id,
            })
        n_leads = self._upsert_batch(rows_leads)

        # -- 3) limpa a flag de quem regularizou -----------------------------
        r = self.sb.table("leads").select("id,name,status,metadata").eq(
            "tenant_id", self.tenant_id
        ).eq("source", "pacto_aluno").not_.is_(
            "metadata->parcela_atrasada", "null"
        ).execute()
        limpos = []
        for lead in (r.data or []):
            if lead["id"] in por_lead:
                continue
            meta = dict(lead.get("metadata") or {})
            meta.pop("parcela_atrasada", None)
            if meta.get("alerta") == "inadimplente":
                meta.pop("alerta", None)
            limpos.append({
                "id":        lead["id"],
                "name":      lead.get("name"),
                "status":    "cliente" if lead.get("status") == "inadimplente" else lead.get("status"),
                "metadata":  meta,
                "tenant_id": self.tenant_id,
            })
        self._upsert_batch(limpos)

        resultado = {
            "alunos_escaneados_com_parcela": len(achados),
            "parcelas_atrasadas":            len(rows_parc),
            "leads_marcados":                n_leads,
            "leads_limpos":                  len(limpos),
        }
        log.info(f"sync_parcelas_atrasadas: {resultado}")
        return resultado

    # -- Prioridade 3: Fases CRM Pacto → sales_pipeline_stages ---------------

    _PIPELINE_NS = uuid.UUID("f1e2d3c4-cafe-5678-abcd-123456789abc")

    def _ensure_pipeline(self, name: str = "CRM Pacto") -> str:
        """Cria ou recupera o pipeline principal. Retorna pipeline_id."""
        pid = str(uuid.uuid5(self._PIPELINE_NS, f"{self.tenant_id}:{name}"))
        self.sb.table("sales_pipelines").upsert({
            "id":         pid,
            "name":       name,
            "description": "Pipeline importado do Pacto Soluções",
            "position":   1,
            "is_default": True,
            "is_active":  True,
            "tenant_id":  self.tenant_id,
        }).execute()
        return pid

    def sync_pipeline_stages_from_pacto(self, adm: "PactoADMClient") -> int:
        """Sincroniza as fases do CRM Pacto para sales_pipeline_stages."""
        log.info("CRM sync: fases do pipeline...")
        fases = adm.fases_crm()
        if not fases:
            log.warning("Nenhuma fase retornada pelo Pacto.")
            return 0

        pipeline_id = self._ensure_pipeline()
        _STAGE_NS   = uuid.UUID("a1b2c3d4-cafe-5678-abcd-fedcba987654")
        cores_padrao = ["#6366f1", "#8b5cf6", "#ec4899", "#f59e0b",
                        "#10b981", "#3b82f6", "#ef4444", "#14b8a6"]

        rows = []
        for i, fase in enumerate(fases):
            name = fase.get("descricao") or fase.get("name") or f"Fase {i+1}"
            slug = fase.get("name", f"fase_{i}")
            sid  = str(uuid.uuid5(_STAGE_NS, f"{pipeline_id}:{slug}"))
            rows.append({
                "id":          sid,
                "name":        name,
                "position":    fase.get("ordem", i + 1),
                "color":       fase.get("cor") or cores_padrao[i % len(cores_padrao)],
                "description": fase.get("name"),
                "is_won":      slug in ("MATRICULA", "RENOVACAO", "REMATRICULA"),
                "is_lost":     slug in ("PERDA", "NAO_CONVERTEU", "OBJECAO"),
                "pipeline_id": pipeline_id,
                "tenant_id":   self.tenant_id,
            })

        for i in range(0, len(rows), 50):
            self.sb.table("sales_pipeline_stages").upsert(rows[i:i+50]).execute()

        log.info(f"sync_pipeline_stages: {len(rows)} fases sincronizadas")
        return len(rows)

    # -- Prioridade 1: Leads de anúncios (Meta Lead Ads) ----------------------

    def sync_meta_lead_ads(self) -> int:
        """
        Migra leads capturados pelo webhook Meta Lead Ads (meta_lead_ads_logs)
        para a tabela leads com utm_source='meta_ads'.
        """
        log.info("CRM sync: Meta Lead Ads → leads...")
        r = self.sb.table("meta_lead_ads_logs").select("*").eq(
            "tenant_id", self.tenant_id
        ).is_("lead_id", "null").execute()
        logs = r.data or []
        if not logs:
            log.info("Nenhum lead Meta Ads pendente de migração.")
            return 0

        rows = []
        for entry in logs:
            lid = entry.get("id")  # usa o próprio id do log como id do lead
            rows.append({
                "id":          lid,
                "name":        entry.get("lead_name"),
                "phone":       "".join(c for c in (entry.get("lead_phone") or "") if c.isdigit()) or None,
                "email":       entry.get("lead_email"),
                "source":      "meta_lead_ads",
                "status":      "lead",
                "utm_source":  "meta_ads",
                "utm_campaign": entry.get("form_name"),
                "tenant_id":   self.tenant_id,
                "metadata": {
                    "meta_page_id":   entry.get("page_id"),
                    "meta_form_id":   entry.get("form_id"),
                    "meta_form_name": entry.get("form_name"),
                    "leadgen_id":     entry.get("leadgen_id"),
                    "raw":            entry.get("raw_data"),
                    "synced_at":      datetime.now().isoformat(),
                },
            })

        n = self._upsert_batch(rows)

        # atualiza lead_id nos logs para não reprocessar
        ids_migrados = [r["id"] for r in rows]
        for i in range(0, len(ids_migrados), 50):
            batch_ids = ids_migrados[i:i+50]
            for log_id in batch_ids:
                self.sb.table("meta_lead_ads_logs").update(
                    {"lead_id": log_id}
                ).eq("id", log_id).execute()

        log.info(f"sync_meta_lead_ads: {n} leads migrados")
        return n

    # -- Prioridade 2: Financeiro → deals --------------------------------------

    _DEAL_NS = uuid.UUID("c0ffee00-cafe-5678-abcd-deadbeef1234")

    def _deal_id(self, pacto_codigo: int, tipo: str) -> str:
        return str(uuid.uuid5(self._DEAL_NS, f"{self.tenant_id}:{pacto_codigo}:{tipo}"))

    def sync_deals_financeiro(self, pacto: "PactoClient", adm: "PactoADMClient") -> dict:
        """
        Cria/atualiza deals para:
          - inadimplentes (contratos vencidos)
          - contratos a vencer em 30 dias
        Usa a fase correta do pipeline se disponível.
        """
        log.info("CRM sync: deals financeiro...")
        pipeline_id = self._ensure_pipeline()

        # tenta mapear stage ids para as fases financeiras
        stages_r = self.sb.table("sales_pipeline_stages").select("id,description").eq(
            "pipeline_id", pipeline_id
        ).execute()
        stage_por_slug = {s["description"]: s["id"] for s in (stages_r.data or [])}
        stage_inadimpl  = stage_por_slug.get("NEGATIVADO") or stage_por_slug.get("PERDA")
        stage_a_vencer  = stage_por_slug.get("RETENCAO")   or stage_por_slug.get("RENOVACAO")

        rows_deals   = []
        rows_leads   = []

        # inadimplentes
        inadimpl = pacto.inadimplentes()
        for a in inadimpl:
            cod  = a.get("codigoCliente")
            if not cod:
                continue
            lead_id = _pacto_lead_id(cod, self.tenant_id)
            rows_leads.append(self._lead_row(cod, a))
            rows_deals.append({
                "id":               self._deal_id(cod, "inadimplente"),
                "lead_id":          lead_id,
                "pipeline_id":      pipeline_id,
                "pipeline_stage_id": stage_inadimpl,
                "title":            f"Inadimplente — {a.get('nome', '')}",
                "original_price":   0.0,
                "status":           "open",
                "payment_status":   "overdue",
                "expected_close_date": (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
                "tenant_id":        self.tenant_id,
                "metadata": {
                    "pacto_codigo":  cod,
                    "dias_atraso":   a.get("dias_atraso"),
                    "vencimento":    a.get("vencimento_fmt"),
                    "synced_at":     datetime.now().isoformat(),
                },
            })

        # contratos a vencer em 30 dias
        a_vencer = pacto.contratos_a_vencer(dias=30)
        codigos_inadimpl = {a.get("codigoCliente") for a in inadimpl}
        for a in a_vencer:
            cod = a.get("codigoCliente")
            if not cod or cod in codigos_inadimpl:
                continue  # já processado como inadimplente
            lead_id = _pacto_lead_id(cod, self.tenant_id)
            rows_leads.append(self._lead_row(cod, a))
            venc_date = (a.get("vencimento_fmt") and
                         datetime.strptime(a["vencimento_fmt"], "%d/%m/%Y").strftime("%Y-%m-%d"))
            rows_deals.append({
                "id":               self._deal_id(cod, "renovacao"),
                "lead_id":          lead_id,
                "pipeline_id":      pipeline_id,
                "pipeline_stage_id": stage_a_vencer,
                "title":            f"Renovação — {a.get('nome', '')}",
                "original_price":   0.0,
                "status":           "open",
                "payment_status":   "pending",
                "expected_close_date": venc_date,
                "tenant_id":        self.tenant_id,
                "metadata": {
                    "pacto_codigo":    cod,
                    "dias_restantes":  a.get("dias_restantes"),
                    "vencimento":      a.get("vencimento_fmt"),
                    "synced_at":       datetime.now().isoformat(),
                },
            })

        n_leads = self._upsert_batch(rows_leads)
        n_deals = 0
        for i in range(0, len(rows_deals), 100):
            r = self.sb.table("deals").upsert(rows_deals[i:i+100]).execute()
            n_deals += len(r.data or [])

        log.info(f"sync_deals_financeiro: {n_leads} leads, {n_deals} deals")
        return {"leads": n_leads, "deals": n_deals}

    # -- Orquestradores por frequência -----------------------------------------

    def sync_visitantes_bv(self, adm: "PactoADMClient", dias: int = 45) -> dict:
        """
        Relatorio 'Conversao de Vendas - BV' por mes (tabela visitantes_bv).

        Cada visitante da fase HO do meta-crm vira uma linha no mes da visita.
        A cada run a situacao acompanha o Pacto: 'Visitante' -> 'Ativo' quando
        um plano e lancado (nunca rebaixa). tipo_bv e derivado UMA vez, no
        primeiro insert, e nunca recalculado:
          - lead inexistente ou ainda visitante/lead  -> Matricula (cadastro novo)
          - lead ja era INATIVO (ex-aluno que voltou) -> Rematricula
          - lead ja cliente/inadimplente (converteu antes do 1o sync): decide
            pela matricula — numero muito abaixo do max da janela = cadastro
            antigo reaproveitado -> Rematricula; senao Matricula.
        Horario roda com dias=2 (novos visitantes); diario com dias=45
        (captura conversoes de quem visitou semanas atras).
        """
        log.info(f"CRM sync: visitantes BV (janela {dias}d)...")
        fim = date.today() + timedelta(days=1)
        ini = fim - timedelta(days=dias)

        # 1. coleta fase HO em janelas de <=30 dias; 1 linha por (cliente, mes)
        vistos: dict[tuple, dict] = {}
        cur = ini
        while cur < fim:
            prox = min(cur + timedelta(days=30), fim)
            codes = []
            for tentativa in (1, 2):  # gateway da 504 esporadico
                try:
                    fases = adm.meta_crm(cur, prox)
                    codes = [c for f in fases for t in f.get("tiposMeta", [])
                             if t.get("faseEnum") == "HO"
                             for c in t.get("codigosFecharMeta", [])]
                    break
                except Exception as e:
                    log.warning(f"meta_crm {cur}..{prox} (tentativa {tentativa}): {e}")
            # ATENCAO: o gateway trunca em 30 linhas e IGNORA page/size
            # (page=1 repete o conteudo da page=0). Fatiar os codigos em
            # lotes pequenos e a unica forma de ler tudo.
            LOTE = 20
            for j in range(0, len(codes), LOTE):
                try:
                    det = adm.meta_crm_detalhada(codes[j:j+LOTE])
                except Exception as e:
                    log.warning(f"meta_crm_detalhada lote {j}: {e} — retry")
                    try:
                        det = adm.meta_crm_detalhada(codes[j:j+LOTE])
                    except Exception as e2:
                        log.warning(f"meta_crm_detalhada lote {j} falhou: {e2}")
                        continue
                for v in det.get("content", []):
                    cod, dm = v.get("codigoCliente"), _data_meta_iso(v.get("dataMeta"))
                    if not cod or not dm:
                        continue
                    # dataMeta = dia do CONTATO de boas-vindas; a visita foi na vespera
                    dv = (date.fromisoformat(dm) - timedelta(days=1)).isoformat()
                    chave = (int(cod), dv[:7])  # mes da visita
                    atual = vistos.get(chave)
                    if not atual or dv > atual["data_visita"]:
                        vistos[chave] = {
                            "codigo_cliente": int(cod),
                            "nome":           v.get("nome"),
                            "matricula":      v.get("matricula"),
                            # detalhada devolve codigo de 2 letras: VI/AT/IN
                            "situacao_pacto": (v.get("situacao") or "").upper(),
                            # Title Case: unifica 'ANDRÉ TREVIZAN' e 'André Trevizan'
                            "consultor":      ((v.get("consultora") or v.get("nomeColaborador") or "").title() or None),
                            "data_visita":    dv,
                        }
            cur = prox
        if not vistos:
            log.info("sync_visitantes_bv: nenhum visitante na janela")
            return {"visitantes": 0, "novos": 0, "convertidos": 0}

        # 2. linhas ja existentes na janela
        meses = sorted({m for _, m in vistos})
        r = self.sb.table("visitantes_bv").select(
            "id,codigo_cliente,mes_referencia,situacao,lead_id"
        ).in_("mes_referencia", meses).execute()
        existentes = {(x["codigo_cliente"], x["mes_referencia"]): x for x in (r.data or [])}

        # 3. leads correspondentes (status decide tipo_bv; id vira FK)
        cods = sorted({c for c, _ in vistos})
        lead_por_cod: dict[int, dict] = {}
        BATCH = 80
        lids = {(_pacto_lead_id(c, self.tenant_id)): c for c in cods}
        todas = list(lids.keys())
        for i in range(0, len(todas), BATCH):
            rr = self.sb.table("leads").select("id,status").in_("id", todas[i:i+BATCH]).execute()
            for row in (rr.data or []):
                lead_por_cod[lids[row["id"]]] = row

        matriculas_num = [int(v["matricula"]) for v in vistos.values()
                          if str(v.get("matricula") or "").isdigit()]
        max_matricula = max(matriculas_num) if matriculas_num else 0

        def _tipo_bv(v: dict) -> str:
            lead = lead_por_cod.get(v["codigo_cliente"])
            status = (lead or {}).get("status")
            if not lead or status in ("lead", "visitante"):
                return "Matricula"
            if status == "inativo":
                return "Rematricula"
            # cliente/inadimplente: converteu antes do 1o sync — decide pela matricula
            m = str(v.get("matricula") or "")
            if m.isdigit() and max_matricula and (max_matricula - int(m)) > 150:
                return "Rematricula"
            return "Matricula"

        agora = datetime.now().isoformat()
        novos, convertidos = [], []
        for (cod, mes), v in vistos.items():
            lead = lead_por_cod.get(cod)
            # convertido = plano lancado no Pacto: a meta diz 'AT' OU o lead ja
            # e cliente na nossa base (cobre metas fechadas, que somem da HO)
            convertido = (v["situacao_pacto"] == "AT"
                          or (lead or {}).get("status") in ("cliente", "inadimplente"))
            ex = existentes.get((cod, mes))
            if ex is None:
                novos.append({
                    "tenant_id":      self.tenant_id,
                    "lead_id":        lead["id"] if lead else None,
                    "codigo_cliente": cod,
                    "nome":           v["nome"],
                    "matricula":      v["matricula"],
                    "situacao":       "Ativo" if convertido else "Visitante",
                    "tipo_bv":        _tipo_bv(v),
                    "consultor":      v["consultor"],
                    "data_visita":    v["data_visita"],
                    "mes_referencia": mes,
                    "convertido_em":  agora if convertido else None,
                })
            elif convertido and ex["situacao"] != "Ativo":
                convertidos.append({"id": ex["id"], "lead_id": ex["lead_id"] or (lead or {}).get("id")})

        # upsert ignorando duplicatas: horario (nuvem) e diario podem se cruzar
        for i in range(0, len(novos), 200):
            self.sb.table("visitantes_bv").upsert(
                novos[i:i+200],
                on_conflict="tenant_id,codigo_cliente,mes_referencia",
                ignore_duplicates=True,
            ).execute()
        for c in convertidos:
            self.sb.table("visitantes_bv").update({
                "situacao": "Ativo", "convertido_em": agora, "lead_id": c["lead_id"],
            }).eq("id", c["id"]).execute()

        # 4. conversoes fora da janela da meta (meta fechada some da fase HO):
        #    qualquer linha ainda 'Visitante' cujo lead virou cliente e promovida
        r = self.sb.table("visitantes_bv").select(
            "id,codigo_cliente").eq("situacao", "Visitante").execute()
        pendentes = r.data or []
        extra = 0
        for i in range(0, len(pendentes), BATCH):
            lote = pendentes[i:i+BATCH]
            lids2 = {_pacto_lead_id(p["codigo_cliente"], self.tenant_id): p for p in lote}
            rr = self.sb.table("leads").select("id,status").in_(
                "id", list(lids2.keys())).execute()
            for row in (rr.data or []):
                if row["status"] in ("cliente", "inadimplente"):
                    self.sb.table("visitantes_bv").update({
                        "situacao": "Ativo", "convertido_em": agora,
                        "lead_id": row["id"],
                    }).eq("id", lids2[row["id"]]["id"]).execute()
                    extra += 1

        # 5. BVs invisiveis a fase HO (incidente 2026-07-08: 49 no relatorio
        #    oficial x 39 na tabela) — meta fechada no MESMO DIA (visitou e
        #    matriculou antes do scan da hora seguinte) ou cadastro sem meta
        #    (ex.: check-in Gympass). Cadastro novo = codigo sequencial novo:
        #    GET /v1/cliente/{codigo} enxerga o cadastro independente da meta.
        #    data_visita = dia em que o lead nasceu no CRM (sync horario de
        #    visitantes cria no mesmo dia) ou hoje.
        recuperados = []
        r = (self.sb.table("visitantes_bv").select("codigo_cliente")
             .order("codigo_cliente", desc=True).limit(1).execute())
        max_cod = int((r.data or [{}])[0].get("codigo_cliente") or 0)
        if max_cod:
            r = self.sb.table("visitantes_bv").select("codigo_cliente").gte(
                "codigo_cliente", max_cod - 80).execute()
            ja_tem = {int(x["codigo_cliente"]) for x in (r.data or [])}
            cod, misses = max_cod - 80, 0
            while misses < 8 and cod < max_cod + 400:
                cod += 1
                if cod in ja_tem:
                    misses = 0
                    continue
                try:
                    c = (adm._gw("GET", f"/v1/cliente/{cod}", timeout=30) or {}).get("content") or {}
                except Exception as e:
                    log.warning(f"bv walker cliente {cod}: {e}")
                    c = {}
                if not c.get("codigo"):
                    if cod > max_cod:
                        misses += 1
                    continue
                misses = 0
                sit = ((c.get("situacao") or {}).get("codigo") or "").upper()
                rl = self.sb.table("leads").select("id,status,created_at").eq(
                    "id", _pacto_lead_id(cod, self.tenant_id)).execute()
                lead = (rl.data or [None])[0]
                dv = ((lead or {}).get("created_at") or date.today().isoformat())[:10]
                # created_at do lead so e proxy confiavel da visita quando o
                # lead nasceu DENTRO da janela (sync horario cria no mesmo
                # dia). Lead antigo = data viria de backfill/import (incidente
                # 2026-07-08: visitantes de junho datados como julho) — pula.
                if lead and date.fromisoformat(dv) < ini:
                    continue
                convertido = (sit == "AT"
                              or (lead or {}).get("status") in ("cliente", "inadimplente"))
                consultor = next((((v.get("colaborador") or {}).get("nome") or "")
                                  for v in (c.get("vinculos") or [])
                                  if v.get("tipoVinculo") == "CONSULTOR"), "")
                recuperados.append({
                    "tenant_id":      self.tenant_id,
                    "lead_id":        (lead or {}).get("id"),
                    "codigo_cliente": cod,
                    "nome":           ((c.get("pessoa") or {}).get("nome")),
                    "matricula":      c.get("matricula"),
                    "situacao":       "Ativo" if convertido else "Visitante",
                    "tipo_bv":        "Rematricula" if (lead or {}).get("status") == "inativo" else "Matricula",
                    "consultor":      (consultor.title() or None),
                    "data_visita":    dv,
                    "mes_referencia": dv[:7],
                    "convertido_em":  agora if convertido else None,
                })

        # 6. rematriculas com codigo ANTIGO que fecharam rapido (meta sumiu da
        #    HO antes do scan): cruza as vendas MA/RE do mes (vendas_pacto) —
        #    venda sem linha BV no mes vira BV 'Ativo' com data da venda.
        mes_atual = date.today().strftime("%Y-%m")
        rv = (self.sb.table("vendas_pacto").select(
                "codigo_cliente,nome_cliente,tipo,data_venda,consultora_vinculo,responsavel_raw")
              .gte("data_venda", f"{mes_atual}-01").in_("tipo", ["MA", "RE"])
              .not_.is_("codigo_cliente", "null").execute())
        rb = self.sb.table("visitantes_bv").select("codigo_cliente").eq(
            "mes_referencia", mes_atual).execute()
        bv_mes = {int(x["codigo_cliente"]) for x in (rb.data or [])}
        bv_mes.update(int(x["codigo_cliente"]) for x in recuperados
                      if x["mes_referencia"] == mes_atual)
        for v in (rv.data or []):
            cod = int(v["codigo_cliente"])
            if cod in bv_mes:
                continue
            bv_mes.add(cod)
            rl = self.sb.table("leads").select("id").eq(
                "id", _pacto_lead_id(cod, self.tenant_id)).execute()
            try:
                c = (adm._gw("GET", f"/v1/cliente/{cod}", timeout=30) or {}).get("content") or {}
            except Exception:
                c = {}
            recuperados.append({
                "tenant_id":      self.tenant_id,
                "lead_id":        ((rl.data or [None])[0] or {}).get("id"),
                "codigo_cliente": cod,
                "nome":           v.get("nome_cliente"),
                "matricula":      c.get("matricula"),
                "situacao":       "Ativo",
                "tipo_bv":        "Rematricula" if v.get("tipo") == "RE" else "Matricula",
                "consultor":      ((v.get("consultora_vinculo") or v.get("responsavel_raw") or "").title() or None),
                "data_visita":    (v.get("data_venda") or date.today().isoformat())[:10],
                "mes_referencia": mes_atual,
                "convertido_em":  agora,
            })

        for i in range(0, len(recuperados), 200):
            self.sb.table("visitantes_bv").upsert(
                recuperados[i:i+200],
                on_conflict="tenant_id,codigo_cliente,mes_referencia",
                ignore_duplicates=True,
            ).execute()

        # 7. auditoria contra o BI oficial "Conversao de Vendas" do ADM
        #    (referencia definida pelo usuario em 2026-07-08): boletinVisitaMes
        #    e o total de BVs do mes — divergencia vira WARNING no log.
        bi_mes = None
        try:
            bi = adm._gw("POST", "/v2-conversao-venda",
                         json={"data": date.today().isoformat()})
            jd = json.loads((bi.get("content") or {}).get("jsonDados", "{}"))
            bi_mes = jd.get("boletinVisitaMes")
        except Exception as e:
            log.warning(f"BI conversao-venda indisponivel: {e}")
        rc = self.sb.table("visitantes_bv").select("id", count="exact").eq(
            "mes_referencia", mes_atual).limit(1).execute()
        tabela_mes = rc.count or 0
        if bi_mes is not None and bi_mes != tabela_mes:
            log.warning(f"visitantes_bv DIVERGE do BI oficial: tabela={tabela_mes} "
                        f"x BI boletinVisitaMes={bi_mes} ({mes_atual})")

        log.info(f"sync_visitantes_bv: {len(vistos)} na janela, "
                 f"{len(novos)} novos, {len(convertidos) + extra} conversoes, "
                 f"{len(recuperados)} recuperados fora da HO, "
                 f"mes {tabela_mes} x BI {bi_mes}")
        return {"visitantes": len(vistos), "novos": len(novos),
                "convertidos": len(convertidos) + extra,
                "recuperados_fora_ho": len(recuperados),
                "tabela_mes": tabela_mes, "bi_boletim_mes": bi_mes}

    def sincronizar_leads(self, pacto: "PactoClient", adm: "PactoADMClient") -> dict:
        """Sync rápido de leads — rodar a cada hora."""
        resultado = {}
        for nome, fn in [
            ("visitantes_hoje", lambda: self.sync_visitantes(adm)),
            # relatorio BV: novos visitantes entram na hora (janela curta)
            ("visitantes_bv_hoje", lambda: self.sync_visitantes_bv(adm, dias=2)),
            ("meta_lead_ads",   lambda: self.sync_meta_lead_ads()),
        ]:
            try:
                resultado[nome] = fn()
            except Exception as e:
                log.error(f"{nome}: {e}")
                resultado[nome] = f"erro: {e}"
        return resultado

    # -- Merge webhook ↔ Pacto ------------------------------------------------

    def merge_leads_webhook_pacto(self) -> int:
        """
        Funde leads criados pelo webhook do WhatsApp (source NULL) com o
        cadastro Pacto da mesma pessoa (mesmos últimos 8 dígitos do fone —
        o mesmo critério de matching que o webhook usa). O lead Pacto
        sobrevive (UUID determinístico, referenciado por deals/parcelas);
        mensagens e métricas de inbox são reapontadas antes de deletar o
        duplicado. Entre múltiplos candidatos Pacto (famílias compartilham
        telefone), prioriza aluno ativo/inadimplente e depois o mais antigo.
        """
        webhook = (self.sb.table("leads")
                   .select("id,name,phone,photo_url")
                   .is_("source", "null").execute().data or [])

        def last8(p) -> str | None:
            d = "".join(c for c in (p or "") if c.isdigit())
            return d[-8:] if len(d) >= 10 else None

        if not any(last8(w.get("phone")) for w in webhook):
            return 0

        # Indexa a base Pacto por últimos 8 dígitos (paginado)
        by8: dict = {}
        page = 0
        while True:
            r = (self.sb.table("leads")
                 .select("id,name,phone,status,photo_url,metadata,created_at")
                 .not_.is_("source", "null")
                 .range(page * 1000, page * 1000 + 999).execute())
            rows = r.data or []
            for p in rows:
                k = last8(p.get("phone"))
                if k:
                    by8.setdefault(k, []).append(p)
            if len(rows) < 1000:
                break
            page += 1

        merged = 0
        for w in webhook:
            k = last8(w.get("phone"))
            candidatos = by8.get(k or "", [])
            if not candidatos:
                continue
            prioridade = {"cliente": 0, "inadimplente": 0}
            keep = sorted(candidatos,
                          key=lambda c: (prioridade.get(c.get("status"), 1),
                                         c.get("created_at") or ""))[0]
            try:
                self.sb.table("whatsapp_messages").update(
                    {"lead_id": keep["id"]}).eq("lead_id", w["id"]).execute()
                # métricas de inbox: reaponta, ou descarta se o Pacto já tem
                tem = (self.sb.table("cs_inbox_metrics").select("id")
                       .eq("lead_id", keep["id"]).limit(1).execute().data or [])
                if tem:
                    self.sb.table("cs_inbox_metrics").delete().eq(
                        "lead_id", w["id"]).execute()
                else:
                    self.sb.table("cs_inbox_metrics").update(
                        {"lead_id": keep["id"]}).eq("lead_id", w["id"]).execute()
                updates = {"metadata": {**(keep.get("metadata") or {}),
                                        "whatsapp_pushname": w.get("name"),
                                        "merged_from_lead": w["id"]}}
                if w.get("photo_url") and not keep.get("photo_url"):
                    updates["photo_url"] = w["photo_url"]
                self.sb.table("leads").update(updates).eq(
                    "id", keep["id"]).execute()
                self.sb.table("leads").delete().eq("id", w["id"]).execute()
                merged += 1
                log.info(f"Merge: '{w.get('name')}' ({w.get('phone')}) → "
                         f"'{keep.get('name')}' [{keep.get('status')}]")
            except Exception as e:
                log.warning(f"Merge falhou {w['id']} → {keep['id']}: {e}")
        log.info(f"Merge webhook↔Pacto: {merged} lead(s) fundido(s)")
        return merged

    def _mapa_datas_visita(self, adm: "PactoADMClient", meses: int = 24) -> dict:
        """
        codigoCliente → data da visita (ISO), via meta-crm fase HO em janelas
        de ~30 dias. Visitantes fora da janela ficam sem data (tratados como
        antigos). ~1 req/mês + páginas do detalhada.
        """
        mapa: dict = {}
        fim = date.today() + timedelta(days=1)
        cur = fim - timedelta(days=meses * 30)
        while cur < fim:
            prox = min(cur + timedelta(days=30), fim)
            try:
                fases = adm.meta_crm(cur, prox)
                codes = [c for f in fases for t in f.get("tiposMeta", [])
                         if t.get("faseEnum") == "HO"
                         for c in t.get("codigosFecharMeta", [])]
                # gateway trunca em 30 linhas e IGNORA page/size — fatiar codigos
                LOTE = 20
                for j in range(0, len(codes), LOTE):
                    det = adm.meta_crm_detalhada(codes[j:j+LOTE])
                    for c in det.get("content", []):
                        cod, dv = c.get("codigoCliente"), _data_meta_iso(c.get("dataMeta"))
                        if cod and dv:
                            # mantém a visita mais RECENTE (repescagens)
                            atual = mapa.get(int(cod))
                            if not atual or dv > atual:
                                mapa[int(cod)] = dv
            except Exception as e:
                log.warning(f"meta_crm {cur}..{prox}: {e}")
            cur = prox
        log.info(f"Mapa de datas de visita: {len(mapa)} visitantes com data")
        return mapa

    def sync_visitantes_antigos(self, pacto: "PactoClient",
                                adm: "PactoADMClient") -> int:
        """
        Todos os cadastros VISITANTE do Pacto viram leads no kanban da aba
        Leads: visita nas últimas 24h → fase 'Visitantes 24h'; passou de
        24h → 'Visitantes Antigos', com metadata.data_visita (ISO) para
        filtrar por data depois. Fases movidas manualmente pelo time para
        FORA dessas duas são preservadas. Chave: codigoCliente.
        """
        log.info("CRM sync: visitantes (base completa, 24h + antigos)...")
        stages = (self.sb.table("sales_pipeline_stages").select("id,name,description")
                  .execute().data or [])
        antigos_id = next((s["id"] for s in stages if s.get("name") == "Visitantes Antigos"), None)
        vinte4_id = next((s["id"] for s in stages
                          if s.get("description") == "VINTE_QUATRO_HORAS"
                          or s.get("name") == "Visitantes 24h"), None)
        if not antigos_id:
            log.warning("Fase 'Visitantes Antigos' não encontrada no pipeline")
            return 0

        mapa_datas = self._mapa_datas_visita(adm)
        corte_24h = (date.today() - timedelta(days=1)).isoformat()

        visitantes = pacto.todos_alunos(situacao="VISITANTE")
        rows = []
        for v in visitantes:
            cod = v.get("codigoCliente")
            if not cod:
                continue
            r = self._lead_row(cod, v, source="pacto_visitante")
            r["status"] = "lead"
            dv = mapa_datas.get(int(cod))
            r["metadata"]["data_visita"] = dv
            r["_stage_alvo"] = (vinte4_id if (dv and dv >= corte_24h and vinte4_id)
                                else antigos_id)
            rows.append(r)

        # preserva fases movidas manualmente para fora do par 24h/antigos
        gerenciadas = {antigos_id, vinte4_id, None}
        preservar = set()
        ids = [r["id"] for r in rows]
        for i in range(0, len(ids), 150):
            got = (self.sb.table("leads").select("id,pipeline_stage_id")
                   .in_("id", ids[i:i + 150]).execute().data or [])
            for g in got:
                if g.get("pipeline_stage_id") not in gerenciadas:
                    preservar.add(g["id"])
        em_24h = em_antigos = 0
        for r in rows:
            alvo = r.pop("_stage_alvo")
            if r["id"] in preservar:
                continue
            r["pipeline_stage_id"] = alvo
            if alvo == vinte4_id:
                em_24h += 1
            else:
                em_antigos += 1

        n = self._upsert_batch(rows)
        log.info(f"sync_visitantes_antigos: {n}/{len(visitantes)} upserted — "
                 f"{em_24h} em Visitantes 24h, {em_antigos} em Antigos, "
                 f"{len(preservar)} fases preservadas")
        return n

    def rotacionar_kanban_leads(self) -> int:
        """
        Move leads do kanban 'Leads Hoje' criados antes de hoje (horário de
        Brasília) para 'Leads Acumuladas'. Os leads entram em 'Leads Hoje'
        pelo webhook do WhatsApp (com origem: anúncio IG/FB, link ou
        espontâneo) e acumulam no dia seguinte se não forem convertidos.
        """
        from datetime import timezone
        stages = (self.sb.table("sales_pipeline_stages").select("id,name")
                  .in_("name", ["Leads Hoje", "Leads Acumuladas"])
                  .execute().data or [])
        por_nome = {s["name"]: s["id"] for s in stages}
        hoje_id = por_nome.get("Leads Hoje")
        acum_id = por_nome.get("Leads Acumuladas")
        if not hoje_id or not acum_id:
            log.warning("Etapas 'Leads Hoje'/'Leads Acumuladas' não encontradas")
            return 0
        tz_sp = timezone(timedelta(hours=-3))
        corte = (datetime.now(tz_sp)
                 .replace(hour=0, minute=0, second=0, microsecond=0)
                 .astimezone(timezone.utc).isoformat())
        r = (self.sb.table("leads")
             .update({"pipeline_stage_id": acum_id})
             .eq("pipeline_stage_id", hoje_id)
             .lt("created_at", corte).execute())
        n = len(r.data or [])
        log.info(f"Kanban leads: {n} movido(s) 'Leads Hoje' → 'Leads Acumuladas'")
        return n

    # nomes canonicos das consultoras (iguais aos de renovacoes/agendamentos)
    _CONSULTORA_CANON = {"kelytta": "Kellyta", "kellyta": "Kellyta",
                         "nathalia": "Nathalia", "nathy": "Nathalia",
                         "raiane": "Raiane", "rai": "Raiane",
                         "ly": "Lyandra", "lyandra": "Lyandra"}

    def sync_vendas_pacto(self, adm: "PactoADMClient", backfill_desde: str | None = None,
                          max_misses: int = 8) -> dict:
        """
        Tabela vendas_pacto: 1 linha por contrato lançado no Pacto (matrícula,
        rematrícula ou renovação) com valor, plano, duração em meses e a
        consultora do lançamento (responsavelLancamento). Alimenta a página
        Ranking de Vendas.

        Códigos de contrato são sequenciais: varre para FRENTE a partir do
        maior código já gravado (para com max_misses buracos seguidos) e, se
        backfill_desde ('YYYY-MM-DD') for dado, também para TRÁS a partir do
        menor código até passar dessa data. Sem linha nenhuma na tabela, o
        seed vem do mapa de matrículas do BI /v2-conversao-venda de hoje.
        """
        from datetime import timezone
        tz_sp = timezone(timedelta(hours=-3))
        log.info("CRM sync: vendas Pacto (contratos)...")

        def _contrato(cod: int) -> dict | None:
            r = adm._gw("GET", f"/contratos/{cod}")
            c = r.get("content") if isinstance(r, dict) else None
            return c if isinstance(c, dict) and c.get("codigo") else None

        def _row(c: dict) -> dict:
            dt = datetime.fromtimestamp((c.get("dataLancamento") or 0) / 1000, tz=tz_sp)
            vig_de = c.get("vigenciaDe")
            vig_ate = c.get("vigenciaAteAjustada") or c.get("vigenciaAte")
            dur = None
            if vig_de and vig_ate:
                dur = max(1, round((vig_ate - vig_de) / 86_400_000 / 30.44))
            resp = (c.get("responsavelLancamento") or "").strip()
            primeiro = resp.split()[0].lower() if resp else ""
            consultora = self._CONSULTORA_CANON.get(primeiro,
                                                    primeiro.title() or None)
            return {
                "tenant_id": self.tenant_id,
                "codigo_contrato": c["codigo"],
                "codigo_pessoa": c.get("pessoa"),
                "nome_cliente": (c.get("pessoaDTO") or {}).get("nome"),
                "tipo": c.get("tipo"),
                "contrato_base": c.get("contratoBaseadoRenovacao") or 0,
                "valor": c.get("valor"),
                "descricao_plano": c.get("descricaoPlano"),
                "vigencia_de": (datetime.fromtimestamp(vig_de / 1000, tz=tz_sp)
                                .date().isoformat() if vig_de else None),
                "vigencia_ate": (datetime.fromtimestamp(vig_ate / 1000, tz=tz_sp)
                                 .date().isoformat() if vig_ate else None),
                "duracao_meses": dur,
                "consultora": consultora,
                "responsavel_raw": resp or None,
                "data_lancamento": dt.astimezone(timezone.utc).isoformat(),
                "data_venda": dt.date().isoformat(),
                "mes_referencia": dt.strftime("%Y-%m"),
                "semana": min((dt.day + 6) // 7, 5),
                "synced_at": datetime.now(timezone.utc).isoformat(),
            }

        r = (self.sb.table("vendas_pacto").select("codigo_contrato")
             .order("codigo_contrato", desc=True).limit(1).execute())
        maior = r.data[0]["codigo_contrato"] if r.data else None

        rows: list[dict] = []
        if maior is None:
            # seed: um contrato de hoje via BI conversão de vendas
            hoje = datetime.now(tz_sp).date().isoformat()
            bi = adm._gw("POST", "/v2-conversao-venda", json={"data": hoje})
            jd = json.loads((bi.get("content") or {}).get("jsonDados", "{}"))
            codigos = (list((jd.get("mapaMatriculasMes") or {}).values())
                       + list((jd.get("mapaRematriculasMes") or {}).values()))
            if not codigos:
                log.warning("sync_vendas_pacto: seed vazio (nenhum contrato no BI)")
                return {"novos": 0}
            maior = max(codigos)
            c = _contrato(maior)
            if c:
                rows.append(_row(c))

        # frente: contratos novos desde o último sync
        cod, misses = maior + 1, 0
        while misses < max_misses:
            c = _contrato(cod)
            if c:
                rows.append(_row(c))
                misses = 0
            else:
                misses += 1
            cod += 1

        # trás: backfill até a data pedida
        if backfill_desde:
            r = (self.sb.table("vendas_pacto").select("codigo_contrato")
                 .order("codigo_contrato").limit(1).execute())
            menor = min([r.data[0]["codigo_contrato"]] if r.data else
                        [x["codigo_contrato"] for x in rows] or [maior])
            cod, misses = menor - 1, 0
            while misses < max_misses:
                c = _contrato(cod)
                if c:
                    row = _row(c)
                    if row["data_venda"] < backfill_desde:
                        break
                    rows.append(row)
                    misses = 0
                else:
                    misses += 1
                cod -= 1

        for i in range(0, len(rows), 100):
            self.sb.table("vendas_pacto").upsert(
                rows[i:i + 100], on_conflict="tenant_id,codigo_contrato").execute()
        log.info(f"sync_vendas_pacto: {len(rows)} contrato(s) gravado(s)")
        return {"novos": len(rows)}

    def sync_vendas_detalhe_mes(self, adm: "PactoADMClient",
                                mes: str | None = None) -> dict:
        """
        Enriquece as vendas do mês (vendas_pacto) para a página Vendas do Mês:
          - codigo_cliente: match do nome_cliente com leads (mesma fonte Pacto)
          - consultora_vinculo: vínculo CONSULTOR do cliente no Pacto (GET
            /v1/cliente/{cod}); se o vínculo mudar no Pacto, atualiza aqui e
            também em leads.metadata.consultora
          - recorrente: regimeRecorrencia do contrato
          - caixa_aberto / parcelas_abertas: soma das parcelas NÃO pagas do
            próprio contrato (GET /parcelas/{codigoPessoa})
        Roda no diário APÓS sync_vendas_pacto. ~3 requests por venda do mês.
        """
        mes = mes or datetime.now().strftime("%Y-%m")
        log.info(f"CRM sync: detalhe das vendas do mês {mes}...")
        r = (self.sb.table("vendas_pacto").select(
                "id,codigo_contrato,codigo_pessoa,nome_cliente,codigo_cliente")
             .eq("mes_referencia", mes).execute())
        vendas = r.data or []
        if not vendas:
            return {"vendas": 0}

        # nome (canônico Pacto) -> codigo_cliente via leads
        nomes = list({v["nome_cliente"] for v in vendas
                      if v.get("nome_cliente") and not v.get("codigo_cliente")})
        nome_para_cod: dict[str, int] = {}
        for i in range(0, len(nomes), 50):
            rl = (self.sb.table("leads").select("name,metadata")
                  .in_("name", nomes[i:i + 50]).execute())
            for ld in rl.data or []:
                cod = (ld.get("metadata") or {}).get("pacto_codigo")
                if cod:
                    nome_para_cod.setdefault(ld["name"], int(cod))

        n = falhas = 0
        for v in vendas:
            upd: dict = {}
            cod_cliente = v.get("codigo_cliente") or nome_para_cod.get(v.get("nome_cliente") or "")
            if cod_cliente and not v.get("codigo_cliente"):
                upd["codigo_cliente"] = cod_cliente

            # vínculo consultora (fonte oficial de atribuição da venda)
            if cod_cliente:
                try:
                    rc = adm._gw("GET", f"/v1/cliente/{cod_cliente}", timeout=30)
                    vincs = (rc.get("content") or {}).get("vinculos") or []
                    nome = next((((x.get("colaborador") or {}).get("nome") or "")
                                 for x in vincs if x.get("tipoVinculo") == "CONSULTOR"), "")
                    if nome.strip():
                        primeiro = nome.strip().split()[0].lower()
                        consultora = self._CONSULTORA_CANON.get(primeiro, primeiro.title())
                        upd["consultora_vinculo"] = consultora
                        # mantém o lead alinhado com o Pacto (pedido do usuário)
                        rl = (self.sb.table("leads").select("id,metadata")
                              .eq("metadata->>pacto_codigo", str(cod_cliente))
                              .limit(1).execute())
                        if rl.data:
                            meta = rl.data[0].get("metadata") or {}
                            if meta.get("consultora") != consultora:
                                meta["consultora"] = consultora
                                (self.sb.table("leads").update({"metadata": meta})
                                 .eq("id", rl.data[0]["id"]).execute())
                except Exception as e:
                    log.warning(f"vinculo venda {v['codigo_contrato']}: {e}")
                    falhas += 1

            # recorrente + caixa em aberto do contrato
            try:
                rc = adm._gw("GET", f"/contratos/{v['codigo_contrato']}")
                c = rc.get("content") or {}
                if isinstance(c, dict) and c.get("codigo"):
                    upd["recorrente"] = bool(c.get("regimeRecorrencia"))
            except Exception as e:
                log.warning(f"contrato {v['codigo_contrato']}: {e}")
                falhas += 1
            try:
                rp = adm._gw("GET", f"/parcelas/{v['codigo_pessoa']}",
                             params={"size": 200})
                parcelas = rp.get("content", []) if isinstance(rp, dict) else []
                abertas = [p for p in parcelas
                           if p.get("contrato") == v["codigo_contrato"]
                           and p.get("situacao") != "PG"]
                upd["caixa_aberto"] = round(sum(float(p.get("valor") or 0)
                                                for p in abertas), 2)
                upd["parcelas_abertas"] = len(abertas)
            except Exception as e:
                log.warning(f"parcelas venda {v['codigo_contrato']}: {e}")
                falhas += 1

            if upd:
                upd["detalhe_synced_at"] = datetime.now(dt_timezone.utc).isoformat()
                self.sb.table("vendas_pacto").update(upd).eq("id", v["id"]).execute()
                n += 1
        log.info(f"sync_vendas_detalhe_mes: {n}/{len(vendas)} vendas ({falhas} falhas)")
        return {"vendas": len(vendas), "atualizadas": n, "falhas": falhas}

    def marcar_recorrentes_parcelas(self, adm: "PactoADMClient") -> int:
        """
        Marca parcelas_atrasadas.recorrente pelo regimeRecorrencia do contrato
        (1 request por CONTRATO distinto sem flag — barato). Alimenta o corte
        'inadimplência de plano recorrente' da página Vendas do Mês.
        """
        r = (self.sb.table("parcelas_atrasadas").select("contrato")
             .is_("recorrente", "null").execute())
        contratos = sorted({x["contrato"] for x in (r.data or []) if x.get("contrato")})
        n = 0
        for cod in contratos:
            try:
                rc = adm._gw("GET", f"/contratos/{cod}")
                c = rc.get("content") or {}
                if isinstance(c, dict) and c.get("codigo"):
                    rec = bool(c.get("regimeRecorrencia"))
                    (self.sb.table("parcelas_atrasadas")
                     .update({"recorrente": rec}).eq("contrato", cod).execute())
                    n += 1
            except Exception as e:
                log.warning(f"recorrente contrato {cod}: {e}")
        log.info(f"marcar_recorrentes_parcelas: {n}/{len(contratos)} contratos")
        return n

    def sync_avaliacoes_fisicas(self, pacto: "PactoClient",
                                mes: str | None = None) -> int:
        """
        Tabela avaliacoes_fisicas_prof: resumo mensal de avaliações físicas
        POR AVALIADOR (login que lançou — pedido do usuário). A listagem
        /psec/avaliacao-fisica-bi/avaliacoes está quebrada no Pacto (500 em
        qualquer formato de filters); a fonte é o resumo
        GET /psec/avaliacao-fisica-bi com codigoAvaliador — validado
        2026-07-07: soma por avaliador == total geral do período.
        1 request por colaborador ativo (~15).
        """
        from datetime import timezone as _tz
        tz_sp = _tz(timedelta(hours=-3))
        agora = datetime.now(tz_sp)
        if mes:
            ano, m = int(mes[:4]), int(mes[5:7])
        else:
            ano, m = agora.year, agora.month
            mes = f"{ano:04d}-{m:02d}"
        ini = datetime(ano, m, 1, tzinfo=tz_sp)
        prox = datetime(ano + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1,
                        tzinfo=tz_sp)
        fim = min(prox - timedelta(seconds=1), agora)

        r = pacto._req("GET", "/psec/colaboradores/all-simple")
        colabs = [c for c in (r.get("content") or [])
                  if c.get("situacao") == "ATIVO"]
        rows: list[dict] = []
        for c in colabs:
            cid = c.get("codigoColaborador") or c.get("id")
            try:
                rb = pacto._req("GET", "/psec/avaliacao-fisica-bi", params={
                    "dataInicio": int(ini.timestamp() * 1000),
                    "dataFim": int(fim.timestamp() * 1000),
                    "codigoAvaliador": cid})
            except Exception as e:
                log.warning(f"avaliacao-fisica-bi avaliador {cid}: {e}")
                continue
            ct = rb.get("content") or {}
            nome = " ".join(w.capitalize() for w in (c.get("nome") or "").split())
            rows.append({
                "tenant_id": self.tenant_id,
                "mes_referencia": mes,
                "professor": nome,
                "codigo_colaborador": cid,
                "avaliacoes": ct.get("avaliacoes") or 0,
                "novas": ct.get("novas") or 0,
                "reavaliacoes": ct.get("reavaliacoes") or 0,
                "synced_at": datetime.now(dt_timezone.utc).isoformat(),
            })
        if rows:
            self.sb.table("avaliacoes_fisicas_prof").upsert(
                rows, on_conflict="tenant_id,mes_referencia,professor").execute()
        total = sum(x["avaliacoes"] for x in rows)
        log.info(f"sync_avaliacoes_fisicas: {mes} — {total} avaliações "
                 f"em {len(rows)} colaboradores")
        return total

    def sync_vendas_avulsas(self, pacto: "PactoClient", adm: "PactoADMClient",
                            mes: str | None = None,
                            base: str = "ativos",
                            pessoa_max: int | None = None) -> dict:
        """
        Tabela vendas_avulsas: produtos de estoque e diárias — parcelas com
        descricao "VENDA AVULSA" e sem contrato, lançadas no mês. Completa o
        Faturamento por período do Pacto (que soma contratos + produto + diária).

        base="ativos": varre as pessoas da base ativa (~2000; roda no diário —
          pega compras de alunos, mas NÃO diárias de visitantes).
        base="completa": varre codigoPessoa 1..pessoa_max (padrão: maior pessoa
          conhecida + 500). ~20k requests, 1-2h — usar para fechar o mês.
        """
        from datetime import timezone as _tz
        tz_sp = _tz(timedelta(hours=-3))
        agora = datetime.now(tz_sp)
        if mes:
            ano, m = int(mes[:4]), int(mes[5:7])
        else:
            ano, m = agora.year, agora.month
            mes = f"{ano:04d}-{m:02d}"
        ini = int(datetime(ano, m, 1, tzinfo=tz_sp).timestamp() * 1000)
        prox = datetime(ano + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1,
                        tzinfo=tz_sp)
        fim = int(prox.timestamp() * 1000)

        if base == "completa":
            if not pessoa_max:
                r = (self.sb.table("vendas_pacto").select("codigo_pessoa")
                     .order("codigo_pessoa", desc=True).limit(1).execute())
                pessoa_max = (r.data[0]["codigo_pessoa"] if r.data else 20500) + 500
            pessoas = list(range(1, pessoa_max + 1))
        else:
            ativos = pacto.alunos_ativos()
            pessoas = sorted({a.get("codigoPessoa") for a in ativos
                              if a.get("codigoPessoa")})
        log.info(f"sync_vendas_avulsas: {mes}, base={base}, "
                 f"{len(pessoas)} pessoas...")

        rows: list[dict] = []
        erros = 0
        for i, cp in enumerate(pessoas):
            try:
                r = adm._gw("GET", f"/parcelas/{cp}", params={"size": 100})
                parcelas = r.get("content", []) if isinstance(r, dict) else []
            except Exception:
                erros += 1
                continue
            for p in parcelas:
                dl = p.get("dataLancamento") or 0
                desc = (p.get("descricao") or "").strip().upper()
                if (ini <= dl < fim and not p.get("contrato")
                        and desc.startswith("VENDA AVULSA")
                        and p.get("situacao") != "RG"):
                    dp = p.get("dataPagamento")
                    rows.append({
                        "tenant_id": self.tenant_id,
                        "parcela_codigo": p.get("codigo"),
                        "codigo_pessoa": cp,
                        "descricao": p.get("descricao"),
                        "valor": p.get("valor"),
                        "situacao": p.get("situacao"),
                        "data_lancamento": datetime.fromtimestamp(
                            dl / 1000, tz=tz_sp).date().isoformat(),
                        "data_pagamento": (datetime.fromtimestamp(
                            dp / 1000, tz=tz_sp).date().isoformat() if dp else None),
                        "mes_referencia": mes,
                        "synced_at": datetime.now(dt_timezone.utc).isoformat(),
                    })
            if (i + 1) % 500 == 0:
                log.info(f"vendas avulsas: {i + 1}/{len(pessoas)} pessoas, "
                         f"{len(rows)} vendas, {erros} erros")
        for i in range(0, len(rows), 100):
            self.sb.table("vendas_avulsas").upsert(
                rows[i:i + 100], on_conflict="tenant_id,parcela_codigo").execute()
        total = sum(float(x["valor"] or 0) for x in rows)
        log.info(f"sync_vendas_avulsas: {len(rows)} vendas avulsas em {mes}, "
                 f"R$ {total:.2f} ({erros} erros)")
        return {"vendas": len(rows), "total": round(total, 2), "erros": erros}

    # -- Relatórios oficiais do ADM (zw-boot) — fonte da página Vendas do Mês --

    def _zwboot(self, pacto: "PactoClient", rota: str, filters: dict,
                timeout: int = 180) -> requests.Response:
        """GET no serviço zw-boot (mesmo backend dos relatórios do ADM novo).
        Auth = Bearer do PactoClient + header empresaId (validado 2026-07-08)."""
        base = pacto.base.split("/TreinoWeb")[0] + "/zw-boot"
        return requests.get(f"{base}{rota}", params={"filters": json.dumps(filters)},
                            headers={"Authorization": f"Bearer {pacto.jwt_token}",
                                     "empresaId": "1"}, timeout=timeout)

    @staticmethod
    def _mes_range_iso(mes: str | None) -> tuple[str, str, str]:
        """('YYYY-MM', dataInicio, dataFinal) na convenção EXATA do SPA do
        ADM: 1º dia e ÚLTIMO dia do mês às 03:00Z (00:00 BRT)."""
        agora = datetime.now()
        mes = mes or agora.strftime("%Y-%m")
        ano, m = int(mes[:4]), int(mes[5:7])
        prox = date(ano + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1)
        ultimo = (prox - timedelta(days=1)).day
        return (mes, f"{ano:04d}-{m:02d}-01T03:00:00.000Z",
                f"{ano:04d}-{m:02d}-{ultimo:02d}T03:00:00.000Z")

    _CANON_CONSULTORA = {"kelytta": "Kellyta", "kellyta": "Kellyta",
                         "nathalia": "Nathalia", "raiane": "Raiane",
                         "lyandra": "Lyandra", "andré": "André", "andre": "André"}

    def sync_comissao_consultora(self, pacto: "PactoClient",
                                 mes: str | None = None) -> dict:
        """
        Relatório oficial "Comissão para Consultor" (Por Faturamento) do ADM →
        tabela comissao_consultora. FONTE OFICIAL do resumo por consultora da
        página Vendas do Mês (pedido do usuário 2026-07-08): planos = contratos
        distintos, valor = soma dos pagamentos, agrupado pela CONSULTORA DO
        CONTRATO (carteira). Receita validada contra a tela do ADM (junho:
        Nathalia 86/89.204,40; Kellyta 58/56.390,80; Raiane 55/55.604,20;
        André 1/1.794,00).
        """
        from io import BytesIO
        from openpyxl import load_workbook

        mes, data_ini, data_fim = self._mes_range_iso(mes)
        log.info(f"CRM sync: comissão p/ consultora {mes} (relatório oficial)...")
        filters = {
            "empresa": "1", "atendente": None, "consultor": None,
            "tipoRelatorioEscolhido": 4,  # Por Faturamento (o que o usuário usa)
            "dataInicio": data_ini, "dataFinal": data_fim,
            "dataInicioR": None, "dataContratosLancadosAPartir": None,
            "dataCompetencia": None, "tipoContrato": None,
            "opcaoImpressao": "CO", "visualizacao": "AP",
            "tipoValorComissoes": "PORC", "tipoResponsavel": 1,
            "retirarRecebiveisComPendecia": False,
            "considerarCompensacaoOriginal": False, "duracoes": [],
        }
        r = self._zwboot(pacto, "/comissao-consultor/gerar-relatorio/excel", filters)
        r.raise_for_status()
        url = (r.json() or {}).get("content")
        if not url:
            raise RuntimeError(f"comissao-consultor sem url de excel: {r.text[:200]}")
        xls = requests.get(url, timeout=180)
        xls.raise_for_status()

        wb = load_workbook(BytesIO(xls.content), data_only=True)
        rows = list(wb.worksheets[0].iter_rows(values_only=True))
        header = [str(c or "").strip() for c in rows[0]]

        def _col(nome: str, fallback: int) -> int:
            # match EXATO primeiro: "Consultor" não pode casar com
            # "Cod. Consultor" (col anterior) — foi o bug que gravou o
            # CÓDIGO da consultora como nome (95 em vez de Nathalia)
            for i, h in enumerate(header):
                if nome.lower() == h.lower().strip():
                    return i
            for i, h in enumerate(header):
                if nome.lower() in h.lower():
                    return i
            return fallback
        i_cons, i_ctr = _col("Consultor", 6), _col("Contrato", 8)
        i_tipo, i_pg = _col("Tipo Contrato", 13), _col("ValorPagamento", 27)

        def _valor(v) -> float:
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v or "").replace("R$", "").replace(".", "").replace(",", ".").strip()
            try:
                return float(s)
            except ValueError:
                return 0.0

        agg: dict[str, dict] = {}
        for linha in rows[1:]:
            if not linha or not linha[i_cons]:
                continue
            bruto = str(linha[i_cons]).strip()
            canon = self._CANON_CONSULTORA.get(
                bruto.split()[0].lower(), " ".join(w.capitalize() for w in bruto.split()[:1]))
            d = agg.setdefault(canon, {"valor": 0.0, "contratos": set(),
                                       "ma": set(), "re": set(), "rn": set()})
            d["valor"] += _valor(linha[i_pg])
            ctr = linha[i_ctr]
            if ctr:
                d["contratos"].add(ctr)
                t = str(linha[i_tipo] or "").strip().upper()
                if t in ("MA", "RE", "RN"):
                    d[t.lower()].add(ctr)

        agora_iso = datetime.now(dt_timezone.utc).isoformat()
        novas = [{
            "tenant_id": self.tenant_id, "mes_referencia": mes,
            "consultora": consultora, "planos": len(d["contratos"]),
            "valor": round(d["valor"], 2),
            "planos_ma": len(d["ma"]), "planos_re": len(d["re"]),
            "planos_rn": len(d["rn"]), "synced_at": agora_iso,
        } for consultora, d in agg.items()]
        # delete+insert: consultora que zerou no mês não fica com linha velha
        self.sb.table("comissao_consultora").delete().eq(
            "mes_referencia", mes).execute()
        if novas:
            self.sb.table("comissao_consultora").insert(novas).execute()
        total = sum(x["valor"] for x in novas)
        log.info(f"sync_comissao_consultora: {mes} — {len(novas)} consultoras, "
                 f"R$ {total:.2f}")
        return {"consultoras": len(novas), "total": round(total, 2)}

    def sync_faturamento_produtos(self, pacto: "PactoClient",
                                  mes: str | None = None) -> dict:
        """
        Relatório oficial "Faturamento por Período" (zw-boot
        /faturamento-sintetico/gerar) → tabela faturamento_produtos: produtos e
        serviços (tudo que NÃO é plano) do mês, geral (consultora NULL) e por
        consultora responsável (param colaboradorCodigo). Alimenta o KPI
        "Produtos e serviços" da página Vendas do Mês.
        """
        mes, data_ini, data_fim = self._mes_range_iso(mes)
        log.info(f"CRM sync: faturamento produtos/serviços {mes}...")
        NAO_PRODUTO = {"mês de referência plano", "mes de referencia plano",
                       "matrícula, rematrícula, renovação",
                       "matricula, rematricula, renovacao",
                       "cheques devolvidos"}
        # cada checkbox de "Tipos de Produtos" da tela é um boolean no filters
        # (nomes descobertos 1 a 1 em 2026-07-08: campo errado dá 500)
        TIPOS_PRODUTO = [
            "manutencaoModalidade", "trancamento", "aulaAvulsa",
            "produtoSessao", "acertoCCAluno", "quitacaoCancelamento",
            "alterarHorario", "bioTotem", "taxaRenegociacao", "produtoEstoque",
            "desafio", "servico", "creditoPersonal", "diaria", "taxaPersonal",
            "pgtoSaldoDevedor", "armario", "consultaNutricional", "locacao",
            "taxaAdesao", "atestado",
        ]

        # códigos das consultoras (colaboradores ativos, match pelo 1º nome)
        rc = pacto._req("GET", "/psec/colaboradores/all-simple")
        consultoras: dict[str, int] = {}
        for c in rc.get("content") or []:
            if c.get("situacao") != "ATIVO":
                continue
            canon = self._CANON_CONSULTORA.get(
                (c.get("nome") or "").split()[0].lower())
            cid = c.get("codigoColaborador") or c.get("id")
            if canon and cid and canon not in consultoras:
                consultoras[canon] = int(cid)

        def _coleta(colab_codigo: int | None, consultora: str | None) -> list[dict]:
            filters = {"empresaId": 1, "dataInicio": data_ini,
                       "dataTermino": data_fim, "agrupamento": "nomeDuracao"}
            for t in TIPOS_PRODUTO:
                filters[t] = True
            if colab_codigo:
                filters["colaboradorCodigo"] = colab_codigo
            r = self._zwboot(pacto, "/faturamento-sintetico/gerar", filters)
            r.raise_for_status()
            content = (r.json() or {}).get("content") or {}
            out = []
            for tp in content.get("listaTipoProduto") or []:
                nome_tipo = (tp.get("tipoProduto") or "").strip()
                if not tp.get("apresentarResultado"):
                    continue
                if nome_tipo.lower() in NAO_PRODUTO:
                    continue
                qtd = valor = 0
                for p in tp.get("listaProduto") or []:
                    # cada tipo tem uma linha 'TOTALIZADOR' que repete a soma
                    # dos produtos — somar junto DOBRA o valor
                    if (p.get("descricao") or "").strip().upper() == "TOTALIZADOR":
                        continue
                    for x in p.get("listaProdutoXMes") or []:
                        qtd += x.get("qtd") or 0
                        valor += x.get("valor") or 0.0
                if qtd or valor:
                    out.append({
                        "tenant_id": self.tenant_id, "mes_referencia": mes,
                        "consultora": consultora, "tipo_produto": nome_tipo,
                        "qtd": qtd, "valor": round(valor, 2),
                        "synced_at": datetime.now(dt_timezone.utc).isoformat(),
                    })
            return out

        rows = _coleta(None, None)
        for consultora, cid in consultoras.items():
            try:
                rows += _coleta(cid, consultora)
            except Exception as e:
                log.warning(f"faturamento produtos {consultora}: {e}")
        self.sb.table("faturamento_produtos").delete().eq(
            "mes_referencia", mes).execute()
        if rows:
            self.sb.table("faturamento_produtos").insert(rows).execute()
        total_geral = sum(x["valor"] for x in rows if x["consultora"] is None)
        log.info(f"sync_faturamento_produtos: {mes} — {len(rows)} linhas, "
                 f"geral R$ {total_geral:.2f}")
        return {"linhas": len(rows), "total_geral": round(total_geral, 2)}

    def sync_parcelas_mes_kpi(self, pacto: "PactoClient",
                              mes: str | None = None) -> dict:
        """
        KPIs "Caixa em aberto (vendas do mês)" e "Inadimplência recorrentes"
        da página Vendas do Mês. Fonte: Relatório Parcelas (zw-boot
        /parcela-em-aberto/consultar), situação EA, faturamento dentro do mês,
        vencimento do dia 01 do mês até o dia 10 do MÊS SEGUINTE; recorrência
        SIM (enum 2) = inadimplência recorrentes, NÃO (enum 3) = caixa aberto.

        REGRA 2026-07-10 (pedido do usuário — caso Denis de Aquino): parcela
        FUTURA não conta. Só entra parcela VENCIDA (dataVencimento < hoje);
        na inadimplência recorrente o aluno ainda precisa ter plano recorrente
        ATIVO (status cliente/inadimplente no CRM). Por isso o cálculo passou
        dos totais prontos do relatório para as linhas, filtradas aqui.
        """
        mes, fat_ini, fat_fim = self._mes_range_iso(mes)
        ano, m = int(mes[:4]), int(mes[5:7])
        prox_ano, prox_m = (ano + 1, 1) if m == 12 else (ano, m + 1)
        venc_ini = fat_ini
        venc_fim = f"{prox_ano:04d}-{prox_m:02d}-10T03:00:00.000Z"
        hoje = date.today().isoformat()
        log.info(f"CRM sync: parcelas em aberto do mês {mes} "
                 f"(somente vencidas < {hoje})...")

        # Matrículas com contrato ativo (cliente/inadimplente) — normalizadas
        # sem zeros à esquerda ("019988" do relatório vs 19988 do sync diário)
        ativos: set[str] = set()
        ini = 0
        while True:
            resp = self.sb.table("leads").select(
                "metadata->>pacto_matricula").in_(
                "status", ["cliente", "inadimplente"]).range(ini, ini + 999).execute()
            rows = resp.data or []
            for row in rows:
                v = str(row.get("pacto_matricula") or "").strip()
                if v.isdigit():
                    ativos.add(str(int(v)))
            if len(rows) < 1000:
                break
            ini += 1000

        def _consulta(recorrencia: int) -> list[dict]:
            filters = {
                "situacoes": ["EA"],
                "dataInicioVencimento": venc_ini,
                "dataTerminoVencimento": venc_fim,
                "dataInicioFaturamento": fat_ini,
                "dataTerminoFaturamento": fat_fim,
                "parcelasRecorrencia": recorrencia,
                "considerarParcelasContratosAssinados": False,
            }
            base = pacto.base.split("/TreinoWeb")[0] + "/zw-boot"
            r = requests.get(f"{base}/parcela-em-aberto/consultar",
                             params={"empresaId": 1, "filters": json.dumps(filters),
                                     "configs": "{}", "page": 0, "size": 500},
                             headers={"Authorization": f"Bearer {pacto.jwt_token}",
                                      "empresaId": "1"}, timeout=90)
            r.raise_for_status()
            c = (r.json() or {}).get("content") or {}
            return c.get("parcelas") or c.get("lista") or []

        def _vencida(p: dict) -> bool:
            return str(p.get("dataVencimento") or "")[:10] < hoje

        def _mat(p: dict) -> str:
            v = str(p.get("matricula") or "").strip()
            return str(int(v)) if v.isdigit() else v

        inad_rows = [p for p in _consulta(2)
                     if _vencida(p) and _mat(p) in ativos]
        caixa_rows = [p for p in _consulta(3) if _vencida(p)]

        inad_valor = round(sum(float(p.get("valor") or 0) for p in inad_rows), 2)
        inad_qtd = len(inad_rows)
        caixa_valor = round(sum(float(p.get("valor") or 0) for p in caixa_rows), 2)
        caixa_qtd = len(caixa_rows)
        self.sb.table("vendas_mes_kpi").upsert({
            "tenant_id": self.tenant_id, "mes_referencia": mes,
            "caixa_aberto_valor": caixa_valor, "caixa_aberto_qtd": caixa_qtd,
            "inad_recorrente_valor": inad_valor, "inad_recorrente_qtd": inad_qtd,
            "synced_at": datetime.now(dt_timezone.utc).isoformat(),
        }, on_conflict="tenant_id,mes_referencia").execute()
        log.info(f"sync_parcelas_mes_kpi: {mes} — caixa {caixa_valor} ({caixa_qtd}) "
                 f"/ inad recorrente {inad_valor} ({inad_qtd})")
        return {"caixa_aberto": caixa_valor, "caixa_qtd": caixa_qtd,
                "inad_recorrente": inad_valor, "inad_qtd": inad_qtd}

    def refresh_instagram_token(self) -> dict:
        """Renova o token da rota 'API com login do Instagram' (config
        INSTAGRAM_LOGIN_TOKEN). Tokens long-lived valem 60 dias; a Meta exige
        idade mínima de 24h pra renovar, então renovamos 1x por semana
        (INSTAGRAM_LOGIN_TOKEN_REFRESHED_AT guarda a última renovação)."""
        r = self.sb.table("config").select("key,value").in_(
            "key", ["INSTAGRAM_LOGIN_TOKEN", "INSTAGRAM_LOGIN_TOKEN_REFRESHED_AT"]).execute()
        cfg = {row["key"]: row["value"] for row in (r.data or [])}
        token = cfg.get("INSTAGRAM_LOGIN_TOKEN")
        if not token:
            return {"status": "sem token configurado"}
        last = cfg.get("INSTAGRAM_LOGIN_TOKEN_REFRESHED_AT")
        if last:
            try:
                idade = datetime.now(dt_timezone.utc) - datetime.fromisoformat(last)
                if idade.days < 7:
                    return {"status": f"renovado há {idade.days}d, pulando"}
            except ValueError:
                pass
        resp = requests.get(
            "https://graph.instagram.com/refresh_access_token",
            params={"grant_type": "ig_refresh_token", "access_token": token},
            timeout=30)
        data = resp.json()
        if resp.status_code != 200 or "access_token" not in data:
            log.error(f"refresh_instagram_token: {data}")
            return {"status": f"erro: {data.get('error', data)}"}
        agora = datetime.now(dt_timezone.utc).isoformat()
        self.sb.table("config").upsert(
            {"key": "INSTAGRAM_LOGIN_TOKEN", "value": data["access_token"]},
            on_conflict="key").execute()
        self.sb.table("config").upsert(
            {"key": "INSTAGRAM_LOGIN_TOKEN_REFRESHED_AT", "value": agora},
            on_conflict="key").execute()
        dias = data.get("expires_in", 0) // 86400
        log.info(f"refresh_instagram_token: renovado, expira em {dias} dias")
        return {"status": "renovado", "expira_em_dias": dias}

    def sincronizar_diario(self, pacto: "PactoClient", adm: "PactoADMClient") -> dict:
        """Sync completo — rodar uma vez ao dia."""
        resultado = {}
        for nome, fn in [
            # renova o token do Instagram (rota login do Instagram) 1x/semana
            ("instagram_token",    lambda: self.refresh_instagram_token()),
            ("alunos_ativos",      lambda: self.sync_alunos_ativos(pacto)),
            # logo após alunos_ativos: quem conversou no WhatsApp antes de
            # virar aluno ganha lead Pacto hoje — funde o lead do webhook nele
            ("merge_webhook_pacto", lambda: self.merge_leads_webhook_pacto()),
            # leads de ontem que não converteram saem de "Leads Hoje"
            ("rotacao_kanban_leads", lambda: self.rotacionar_kanban_leads()),
            # base histórica de visitantes → fase "Visitantes Antigos"
            ("visitantes_antigos",  lambda: self.sync_visitantes_antigos(pacto, adm)),
            # relatorio BV: rescan 45d captura conversoes Visitante -> Ativo
            ("visitantes_bv",      lambda: self.sync_visitantes_bv(adm, dias=45)),
            # depois de alunos_ativos (que reescreve o metadata) pra re-marcar a flag
            ("aniversariantes",    lambda: self.sync_aniversariantes(pacto)),
            ("grupo_risco",        lambda: self.sync_grupo_risco(adm)),
            ("pipeline_stages",    lambda: self.sync_pipeline_stages_from_pacto(adm)),
            ("deals_financeiro",   lambda: self.sync_deals_financeiro(pacto, adm)),
            ("renovacao_mes_atual", lambda: self.sync_renovacao_mes_atual(pacto)),
            # depois de alunos_ativos (metadata.vencimento fresco): marca
            # 'renovado' na pagina Renovacao quem lancou contrato novo no Pacto
            ("renovacoes_status",  lambda: self.sync_renovacoes_status()),
            # idem pra pagina Agendamentos: quem agendou aula e virou aluno
            ("agendamentos_status", lambda: self.sync_agendamentos_status()),
            # e pra pagina Leads (acompanhamento)
            ("leads_acomp_status", lambda: self.sync_leads_acompanhamento_status()),
            # contratos lançados (matrícula/rematrícula/renovação) → Ranking
            ("vendas_pacto",       lambda: self.sync_vendas_pacto(adm)),
            # página Vendas do Mês: vínculo consultora + caixa aberto + recorrente
            ("vendas_detalhe_mes", lambda: self.sync_vendas_detalhe_mes(adm)),
            # página Ranking Profs: avaliações físicas por professor
            ("avaliacoes_fisicas", lambda: self.sync_avaliacoes_fisicas(pacto)),
            # página Vendas do Mês: relatórios OFICIAIS do ADM (zw-boot) —
            # comissão por consultora (carteira) + produtos/serviços
            ("comissao_consultora", lambda: self.sync_comissao_consultora(pacto)),
            ("faturamento_produtos", lambda: self.sync_faturamento_produtos(pacto)),
            ("parcelas_mes_kpi",   lambda: self.sync_parcelas_mes_kpi(pacto)),
            # mês ANTERIOR também: o caixa em aberto dele continua vivo (parcelas
            # de junho seguem sendo pagas em julho) — sem isso a página Vendas do
            # Mês mostra uma foto congelada do mês fechado (bug achado 13/07/26:
            # junho travado em 13.526,65 desde 09/07 vs 6.109,45 real no Pacto)
            ("parcelas_mes_kpi_ant", lambda: self.sync_parcelas_mes_kpi(
                pacto, mes=(date.today().replace(day=1) - timedelta(days=1)).strftime("%Y-%m"))),
            ("inadimplentes",      lambda: self.sync_inadimplentes(pacto, adm)),
            # base completa (~2000, ~45min): roda de madrugada junto do diario;
            # achou 22% mais parcelas que o subset de risco (medido 2026-07-02)
            ("parcelas_atrasadas", lambda: self.sync_parcelas_atrasadas(pacto, adm, todos_ativos=True)),
            # depois de alunos_ativos (metadata zerado) e parcelas (lista alvo
            # fresca): consultora do vinculo pro resumo de Inadimplencia
            ("consultora_vinculo", lambda: self.sync_consultora_vinculo(adm)),
            # depois de parcelas_atrasadas: flag recorrente por contrato
            ("recorrentes_parcelas", lambda: self.marcar_recorrentes_parcelas(adm)),
            # 1 request/aluno (~15min na base completa) — alimenta o card do
            # Kanban de Alunos ("Xd sem vir")
            ("ultimo_acesso",      lambda: self.sync_ultimo_acesso(pacto)),
            # produtos/diárias comprados pela base ativa (visitantes: varredura
            # completa mensal via menu 54 base=completa). É o passo mais lento
            # (~75 min) e o menos crítico — fica por ÚLTIMO de propósito: se o
            # diário estourar o timeout de novo, perde-se só este KPI e não o
            # enriquecimento de metadata (incidentes 2026-07-06 e 2026-07-08)
            ("vendas_avulsas",     lambda: self.sync_vendas_avulsas(pacto, adm)),
        ]:
            try:
                resultado[nome] = fn()
            except Exception as e:
                log.error(f"{nome}: {e}")
                resultado[nome] = f"erro: {e}"
        return resultado

    def sincronizar_tudo(self, pacto: "PactoClient", adm: "PactoADMClient") -> dict:
        """Executa todos os syncs (leads + diário) e retorna resumo consolidado."""
        log.info("=== Sincronização completa Pacto → CRM ===")
        r1 = self.sincronizar_leads(pacto, adm)
        r2 = self.sincronizar_diario(pacto, adm)
        resultado = {**r1, **r2}
        log.info(f"=== Resultado: {resultado} ===")
        return resultado


def crm_run_scheduler(crm: "CRMClient", pacto: PactoClient,
                      adm: PactoADMClient) -> None:
    """
    Inicia o scheduler de sincronização contínua:
      - a cada hora: visitantes + Meta Lead Ads
      - todos os dias às 07:00: alunos, grupo de risco, fases, deals financeiro
    Mantém o processo rodando até Ctrl+C.
    """
    try:
        import schedule
    except ImportError:
        raise RuntimeError("Pacote 'schedule' não instalado. Execute: pip install schedule")

    log.info("Scheduler CRM iniciado (Ctrl+C para parar).")

    schedule.every().hour.do(crm.sincronizar_leads, pacto, adm)
    schedule.every().day.at("07:00").do(crm.sincronizar_diario, pacto, adm)

    # executa imediatamente na primeira rodada
    crm.sincronizar_tudo(pacto, adm)

    while True:
        schedule.run_pending()
        time.sleep(60)


# -- CLI interativo -----------------------------------------------------------

if __name__ == "__main__":
    import sys

    pacto = PactoClient()
    adm   = PactoADMClient()
    crm   = None  # inicializado sob demanda (requer supabase instalado)

    # Modos nao-interativos (GitHub Actions / Task Scheduler):
    #   python agente_integrador_pacto.py --scan-parcelas-completo  (so parcelas, base inteira)
    #   python agente_integrador_pacto.py --sync-diario             (sincronizar_tudo)
    #   python agente_integrador_pacto.py --sync-horario            (sincronizar_leads)
    if "--sync-horario" in sys.argv:
        r = CRMClient().sincronizar_leads(pacto, adm)
        print(json.dumps(r, ensure_ascii=False, indent=2, default=str))
        raise SystemExit(1 if any(isinstance(v, str) and v.startswith("erro") for v in r.values()) else 0)
    if "--scan-parcelas-completo" in sys.argv:
        r = CRMClient().sync_parcelas_atrasadas(pacto, adm, todos_ativos=True)
        print(json.dumps(r, ensure_ascii=False, indent=2))
        raise SystemExit(0)
    if "--sync-diario" in sys.argv:
        r = CRMClient().sincronizar_tudo(pacto, adm)
        print(json.dumps(r, ensure_ascii=False, indent=2, default=str))
        # falha o job se algum sync deu erro (visivel no historico do Actions)
        raise SystemExit(1 if any(isinstance(v, str) and v.startswith("erro") for v in r.values()) else 0)

    def _crm() -> CRMClient:
        global crm
        if crm is None:
            crm = CRMClient()
        return crm

    MENU = """
+------------------------------------------+
|  Agente Integrador Pacto - Territorio Fit |
+------------------------------------------+
 --- TreinoWeb ---
  1 - Alunos ativos
  2 - Alunos faltosos (7 dias)
  3 - Alunos faltosos (14 dias)
  4 - Inadimplentes (contrato vencido)
  5 - Contratos a vencer (30 dias)
  6 - Cancelamentos (30 dias)
  7 - Renovacoes ativas
  8 - Matriculas ativas
  9 - Alunos inativos
 10 - Checkins recentes
 11 - Dashboard completo + exportar MD
 12 - Mapear endpoints
 13 - Usuario autenticado
 --- CRM Pacto (ADM Gateway) ---
 14 - Meta de hoje (abertura + visitantes 24h)
 15 - Fases do funil CRM
 16 - Objecoes cadastradas
 17 - Meta diaria (ultimos 7 dias)
 --- Financeiro (ADM Gateway) ---
 18 - Resumo financeiro (movcontas)
 19 - Recebiveis (tipos 3+8)
 20 - Despesas (tipo 1)
 21 - Grupo de risco (churn)
 22 - Planos cadastrados
 --- Presenca / Acesso Fisico ---
 23 - Lista acessos recentes (academia, hoje)
 24 - Historico presenca aluno (por matricula)
 25 - Distribuicao semanal de acesso fisico (por cod_aluno)
 26 - Reconstruir peso de risco (por cod_aluno)
 --- CRM Supabase (crm.territoriofit.com.br) ---
 27 - Sync: alunos ativos -> leads CRM
 28 - Sync: visitantes hoje -> leads CRM
 29 - Sync: grupo de risco -> leads CRM (metadata)
 30 - Sync: fases Pacto -> sales_pipeline_stages
 31 - Sync: Meta Lead Ads -> leads CRM
 32 - Sync: inadimplentes + a vencer -> deals CRM
 33 - Sync completo (todos os grupos)
 34 - Iniciar scheduler (horario + diario, loop continuo)
 35 - Sync: alunos inativos -> leads CRM
 36 - Sync: renovacao mes atual -> etapa Renovacao CRM
 37 - Sync: inadimplentes (parcelas em aberto) -> etapa Inadimplente CRM
 38 - Sync: parcelas atrasadas EA (risco peso>=5 + inadimplentes, ~740)
 39 - Sync: parcelas atrasadas EA (TODOS os ativos, ~2000 -- lento)
 40 - Sync: ultimo acesso catraca -> leads CRM (metadata, ~15min)
 41 - Sync: aniversariantes do mes -> leads CRM (metadata)
 42 - Merge: leads do webhook WhatsApp -> cadastro Pacto (mesmo fone)
 43 - Kanban: mover 'Leads Hoje' de ontem -> 'Leads Acumuladas'
 44 - Sync: visitantes antigos (base completa) -> fase Visitantes Antigos
 45 - Sync: detectar renovacoes lancadas no Pacto -> pagina Renovacao
 46 - Sync: visitantes BV (conversao de vendas) -> tabela visitantes_bv
 47 - Sync: detectar fechamentos no Pacto -> pagina Agendamentos
 48 - Sync: consultora do vinculo (inadimplentes) -> leads.metadata
 49 - Sync: detectar fechamentos no Pacto -> pagina Leads (acompanhamento)
 50 - Sync: vendas Pacto (contratos) -> tabela vendas_pacto (Ranking)
 51 - Sync: detalhe vendas do mes (vinculo/caixa aberto/recorrente)
 52 - Sync: flag recorrente nas parcelas atrasadas
 53 - Sync: avaliacoes fisicas (BI) -> tabela avaliacoes_fisicas
 54 - Sync: vendas avulsas (produto/diaria) -> tabela vendas_avulsas
  0 - Sair
"""

    while True:
        print(MENU)
        op = input("Escolha: ").strip()

        if op == "0":
            break
        elif op == "1":
            r = pacto.alunos_ativos()
            print(f"\nAtivos: {len(r)}")
            pacto.exportar("alunos_ativos", r)
        elif op == "2":
            r = pacto.alunos_faltosos(dias=7)
            print(f"\nFaltosos +7d: {len(r)}")
            for a in r[:10]:
                print(f"  {a.get('nome')[:35]:35} | {a.get('dias_ausente')}d | {a.get('ultimo_treino') or 'nenhum'}")
            pacto.exportar("faltosos_7d", r)
        elif op == "3":
            r = pacto.alunos_faltosos(dias=14)
            print(f"\nFaltosos +14d: {len(r)}")
            for a in r[:10]:
                print(f"  {a.get('nome')[:35]:35} | {a.get('dias_ausente')}d")
            pacto.exportar("faltosos_14d", r)
        elif op == "4":
            r = pacto.inadimplentes()
            print(f"\nInadimplentes: {len(r)}")
            for a in r[:10]:
                print(f"  {a.get('nome')[:35]:35} | venc: {a.get('vencimento_fmt')} | {a.get('dias_atraso')}d atraso")
            pacto.exportar("inadimplentes", r)
        elif op == "5":
            r = pacto.contratos_a_vencer()
            print(f"\nA vencer em 30d: {len(r)}")
            for a in r[:10]:
                print(f"  {a.get('nome')[:35]:35} | vence: {a.get('vencimento_fmt')} | {a.get('dias_restantes')}d")
            pacto.exportar("a_vencer", r)
        elif op == "6":
            r = pacto.cancelamentos()
            print(f"\nCancelamentos (30d): {len(r)}")
            pacto.exportar("cancelamentos", r)
        elif op == "7":
            r = pacto.renovacoes()
            print(f"\nRenovacoes ativas: {len(r)}")
            pacto.exportar("renovacoes", r)
        elif op == "8":
            r = pacto.matriculas_novas()
            print(f"\nMatriculas ativas: {len(r)}")
            pacto.exportar("matriculas", r)
        elif op == "9":
            r = pacto.alunos_inativos()
            print(f"\nInativos: {len(r)}")
            pacto.exportar("inativos", r)
        elif op == "10":
            r = pacto.checkins_recentes()
            print(f"\nCheckins recentes: {len(r)}")
            for c in r[:10]:
                print(f"  {c.get('nome')[:30]:30} | {c.get('dataHora')} | {c.get('tipoCheckin')}")
        elif op == "11":
            painel = pacto.dashboard()
            pacto.exportar("dashboard", painel)
            md = pacto.exportar_dashboard_md(painel)
            print(f"\nDashboard:")
            print(f"   Ativos:          {painel['alunos']['ativos']}")
            print(f"   Faltosos +7d:    {painel['alunos']['faltosos_7_dias']}")
            print(f"   Faltosos +14d:   {painel['alunos']['faltosos_14_dias']}")
            print(f"   Inadimplentes:   {painel['contratos']['inadimplentes']}")
            print(f"   A vencer 30d:    {painel['contratos']['a_vencer_30_dias']}")
            print(f"   Cancelamentos:   {painel['contratos']['cancelamentos_30d']}")
            print(f"\n   MD: {md}")
        elif op == "12":
            r = pacto.listar_endpoints()
            print(json.dumps(r, ensure_ascii=False, indent=2)[:3000])
            pacto.exportar("endpoints", r)
        elif op == "13":
            r = pacto.obter_usuario()
            u = r.get("user", {})
            print(f"\nUsuario: {u.get('nome')} (@{u.get('username')})")
            print(f"Perfil:  {', '.join(u.get('perfis', []))}")
            emp = r.get("unidadesEmpresa", [{}])[0]
            print(f"Empresa: {emp.get('nome')} (id={emp.get('id')})")
            recursos = r.get("perfilUsuario", {}).get("recursos", [])
            print(f"Recursos: {len(recursos)} permissoes")
            funcs = [f for f in r.get("perfilUsuario", {}).get("funcionalidades", []) if f.get("possuiFuncionalidade")]
            print(f"Funcionalidades ativas: {len(funcs)}")
            pacto.exportar("usuario", r)
        elif op == "14":
            ab = adm.meta_crm_abertura()
            print(f"\nMeta hoje: abriu={ab.get('abriuMetaHoje')} | pode abrir={ab.get('metaPodeSerAberta')}")
            if ab.get("mensagemBloqueio"):
                print(f"Bloqueio: {ab.get('mensagemBloqueio')}")
            v = adm.visitantes_24h(date.today())
            print(f"Visitantes 24h: {v.get('total')} total | meta={v.get('meta')} | realizado={v.get('realizado')}")
            for vis in v.get("visitantes", [])[:10]:
                print(f"  {(vis.get('nome') or '?')[:35]:35} | {vis.get('situacao')} | {vis.get('consultora') or '—'}")
            pacto.exportar("visitantes_24h", v)
        elif op == "15":
            fases = adm.fases_crm()
            print(f"\nFases CRM ({len(fases)} total):")
            for f in fases:
                print(f"  {f.get('name','?'):35} {f.get('descricao','')[:40]}")
        elif op == "16":
            obj = adm.objecoes()
            print(f"\nObjecoes ({len(obj)}):")
            for o in obj:
                ativo = "ativo" if o.get("ativo") else "inativo"
                print(f"  [{o.get('codigo')}] {o.get('descricao','?')} ({ativo})")
        elif op == "17":
            hoje = date.today()
            inicio = hoje - timedelta(days=7)
            r = adm.meta_diaria(inicio, hoje)
            print(f"\nMeta diaria ({inicio} a {hoje}): {len(r)} registros")
            for m in r[:10]:
                print(f"  {json.dumps(m, ensure_ascii=False)[:100]}")
            if r:
                pacto.exportar("meta_diaria", r)
        elif op == "18":
            r = adm.resumo_financeiro()
            print(f"\nResumo financeiro:")
            print(f"  Lancamentos: {r.get('total_lancamentos')}")
            print(f"  Recebiveis:  R$ {r.get('recebiveis_total'):,.2f}")
            print(f"  Despesas:    R$ {r.get('despesas_total'):,.2f}")
            print(f"  Estornos:    R$ {r.get('estornos_total'):,.2f}")
            print(f"  Saldo:       R$ {r.get('saldo'):,.2f}")
            print(f"  Por tipo:    {r.get('por_tipo')}")
            pacto.exportar("resumo_financeiro", r)
        elif op == "19":
            r = adm.receitas()
            print(f"\nRecebiveis (tipos 3+8): {len(r)}")
            for m in r[:10]:
                print(f"  {m.get('descricao','?')[:45]:45} | R$ {m.get('valor')} | {m.get('dataLancamento','')[:10]}")
            pacto.exportar("recebiveis", r)
        elif op == "20":
            r = adm.despesas()
            print(f"\nDespesas (tipo 1): {len(r)}")
            for m in r[:10]:
                print(f"  {m.get('descricao','?')[:45]:45} | R$ {m.get('valor')} | {m.get('dataLancamento','')[:10]}")
            pacto.exportar("despesas", r)
        elif op == "21":
            r = adm.grupo_risco()
            print(f"\nGrupo de risco (churn): {len(r)} clientes")
            for c in r[:10]:
                print(f"  [{c.get('cliente')}] {c.get('nomeCliente','?')[:40]}")
            pacto.exportar("grupo_risco", r)
        elif op == "22":
            r = adm.planos()
            print(f"\nPlanos ({len(r)}):")
            for p in r[:15]:
                print(f"  [{p.get('codigo')}] {p.get('descricao','?')[:50]}")
            pacto.exportar("planos", r)
        elif op == "23":
            r = pacto.lista_acessos_recentes(tipo=1, limite=50)
            print(f"\nAcessos fisicos recentes: {len(r)}")
            for a in r[:15]:
                print(f"  [{a.get('matricula')}] {a.get('nome','?')[:35]:35} | {a.get('hora','')} | {a.get('plano','')[:25]}")
        elif op == "24":
            mat = int(input("Matricula ZW do aluno: ").strip())
            r = pacto.historico_presenca(mat)
            print(f"\nHistorico de presenca (mat={mat}):")
            print(f"  Total aulas realizadas : {r.get('totalAulasRealizadas','?')}")
            print(f"  Aulas no mes atual     : {r.get('aulasMesAtual','?')}")
            print(f"  Semanas consecutivas   : {r.get('semanasConsecutivas','?')}")
        elif op == "25":
            cod = int(input("Codigo do aluno (codigoCliente): ").strip())
            r = pacto.distribuicao_acesso_semanal(cod)
            total = sum(r.values()) if r else 0
            print(f"\nDistribuicao de acesso fisico semanal (cod={cod}, ultimos 6 meses):")
            for dia, cnt in r.items():
                bar = "#" * cnt
                print(f"  {dia}: {cnt:3}  {bar}")
            print(f"  Total: {total} acessos")
        elif op == "26":
            cod = int(input("Codigo do aluno (codigoCliente): ").strip())
            r = adm.reconstruir_peso(cod)
            print(f"\nPeso de risco reconstruido (cod={cod}):")
            print(f"  Peso calculado : {r['peso_calculado']}")
            print(f"  Vencimento     : {r['peso_vencimento']} ({r['vencimento']}"
                  + (f", {r['dias_para_vencer']}d" if r['dias_para_vencer'] is not None else "") + ")")
            print(f"  Faltas         : {r['peso_faltas']} ({r['dias_ausente_aprox']} sem acesso)")
            print(f"  Presenca 4 sem : {r['peso_presenca']} ({r['media_presenca_4sem']:.1f} dias/sem)")
            print(f"  Formula        : {r['detalhes']}")
        elif op == "27":
            n = _crm().sync_alunos_ativos(pacto)
            print(f"\nSync alunos ativos → CRM: {n} leads upserted")
        elif op == "28":
            n = _crm().sync_visitantes(adm)
            print(f"\nSync visitantes hoje → CRM: {n} leads upserted")
        elif op == "29":
            n = _crm().sync_grupo_risco(adm)
            print(f"\nSync grupo de risco → CRM: {n} leads atualizados")
        elif op == "30":
            n = _crm().sync_pipeline_stages_from_pacto(adm)
            print(f"\nSync fases CRM → sales_pipeline_stages: {n} fases")
        elif op == "31":
            n = _crm().sync_meta_lead_ads()
            print(f"\nSync Meta Lead Ads → leads CRM: {n} leads migrados")
        elif op == "32":
            r = _crm().sync_deals_financeiro(pacto, adm)
            print(f"\nSync financeiro → deals CRM: {r['leads']} leads, {r['deals']} deals")
        elif op == "33":
            r = _crm().sincronizar_tudo(pacto, adm)
            print("\nSync completo concluido:")
            for k, v in r.items():
                print(f"  {k:25}: {v}")
        elif op == "34":
            print("\nIniciando scheduler (Ctrl+C para parar)...")
            crm_run_scheduler(_crm(), pacto, adm)
        elif op == "35":
            n = _crm().sync_alunos_inativos(pacto)
            print(f"\nSync alunos inativos → CRM: {n} leads upserted")
        elif op == "36":
            n = _crm().sync_renovacao_mes_atual(pacto)
            print(f"\nSync renovação mês atual → CRM: {n} leads na etapa Renovação")
        elif op == "37":
            n = _crm().sync_inadimplentes(pacto, adm)
            print(f"\nSync inadimplentes → CRM: {n} leads (com dados de parcela em aberto)")
        elif op == "38":
            r = _crm().sync_parcelas_atrasadas(pacto, adm)
            print(f"\nSync parcelas atrasadas → CRM:")
            for k, v in r.items():
                print(f"  {k:32}: {v}")
        elif op == "39":
            r = _crm().sync_parcelas_atrasadas(pacto, adm, todos_ativos=True)
            print(f"\nSync parcelas atrasadas (base completa) → CRM:")
            for k, v in r.items():
                print(f"  {k:32}: {v}")
        elif op == "40":
            n = _crm().sync_ultimo_acesso(pacto)
            print(f"\nSync ultimo acesso catraca → CRM: {n} leads atualizados")
        elif op == "41":
            n = _crm().sync_aniversariantes(pacto)
            print(f"\nSync aniversariantes do mes → CRM: {n} leads atualizados")
        elif op == "42":
            n = _crm().merge_leads_webhook_pacto()
            print(f"\nMerge leads webhook ↔ Pacto: {n} lead(s) fundido(s)")
        elif op == "43":
            n = _crm().rotacionar_kanban_leads()
            print(f"\nKanban: {n} lead(s) movido(s) para 'Leads Acumuladas'")
        elif op == "44":
            n = _crm().sync_visitantes_antigos(pacto, adm)
            print(f"\nSync visitantes antigos → CRM: {n} leads")
        elif op == "45":
            r = _crm().sync_renovacoes_status()
            print(f"\nRenovações detectadas no Pacto: {r['renovadas']} de {r['abertas']} abertas")
        elif op == "46":
            dias = input("Janela em dias [45]: ").strip()
            r = _crm().sync_visitantes_bv(adm, dias=int(dias) if dias else 45)
            print(f"\nVisitantes BV: {r}")
        elif op == "47":
            r = _crm().sync_agendamentos_status()
            print(f"\nAgendamentos: {r['fechados']} fechamento(s) detectado(s), "
                  f"{r['vinculados']} vinculo(s) novo(s), {r['abertos']} aberto(s)")
        elif op == "48":
            n = _crm().sync_consultora_vinculo(adm)
            print(f"\nConsultora do vinculo: {n} lead(s) atualizados")
        elif op == "49":
            r = _crm().sync_leads_acompanhamento_status()
            print(f"\nLeads acompanhamento: {r['fechados']} fechado(s), "
                  f"{r['vinculados']} vinculo(s) novo(s), {r['abertos']} aberto(s)")
        elif op == "50":
            desde = input("Backfill desde (YYYY-MM-DD, vazio = só novos): ").strip() or None
            r = _crm().sync_vendas_pacto(adm, backfill_desde=desde)
            print(f"\nVendas Pacto: {r['novos']} contrato(s) gravado(s)")
        elif op == "51":
            mes = input("Mês (YYYY-MM, vazio = atual): ").strip() or None
            r = _crm().sync_vendas_detalhe_mes(adm, mes=mes)
            print(f"\nDetalhe vendas: {r}")
        elif op == "52":
            n = _crm().marcar_recorrentes_parcelas(adm)
            print(f"\nRecorrentes: {n} contrato(s) marcados")
        elif op == "53":
            n = _crm().sync_avaliacoes_fisicas(pacto)
            print(f"\nAvaliações físicas: {n} registro(s)")
        elif op == "54":
            mes = input("Mês (YYYY-MM, vazio = atual): ").strip() or None
            base = input("Base (ativos/completa, vazio = ativos): ").strip() or "ativos"
            r = _crm().sync_vendas_avulsas(pacto, adm, mes=mes, base=base)
            print(f"\nVendas avulsas: {r}")
        elif op == "55":
            mes = input("Mês (YYYY-MM, vazio = atual): ").strip() or None
            r = _crm().sync_comissao_consultora(pacto, mes=mes)
            print(f"\nComissão consultora: {r}")
        elif op == "56":
            mes = input("Mês (YYYY-MM, vazio = atual): ").strip() or None
            r = _crm().sync_faturamento_produtos(pacto, mes=mes)
            print(f"\nFaturamento produtos: {r}")
        elif op == "57":
            mes = input("Mês (YYYY-MM, vazio = atual): ").strip() or None
            r = _crm().sync_parcelas_mes_kpi(pacto, mes=mes)
            print(f"\nParcelas do mês (caixa/inad): {r}")
        else:
            print("Opcao invalida.")
